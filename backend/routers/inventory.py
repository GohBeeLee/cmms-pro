"""
Inventory Router
================
Uses real model fields: part_code, name, category, description,
quantity_on_hand, reorder_level, unit_cost, supplier, location,
barcode, used_on_asset, notes, unit.

IMPORTANT: Static routes (/import/*, /export/*) declared BEFORE
           /{part_id} routes to avoid FastAPI path conflicts.
"""
import io
import re
from datetime import datetime, timedelta
from typing import Optional
from uuid import UUID

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from pydantic import BaseModel

from db import get_db
from models import SparePart
from auth import get_current_user
from models import User
from websocket_manager import ws_manager

router = APIRouter(prefix="/inventory", tags=["inventory"])
ROOM  = "inventory"

# Default spare-part categories — always offered in dropdowns/filters even
# before any part exists in them yet. Categories found in an Excel import
# that don't match one of these are NOT discarded or forced into "Other" —
# they're kept exactly as written and simply become additional categories.
DEFAULT_CATEGORIES = [
    "Bearing","Belt and Chain","Filter","Electrical Components",
    "Electronic Components","Wear Parts","Fastener and Hardware",
    "Lubricants and Grease","Pneumatic Components","Pneumatic Fitting",
    "Sensor and Instrumentation",
]


# ── Schema ─────────────────────────────────────────────────────────────────

class PartIn(BaseModel):
    part_code:        Optional[str]   = None
    name:             Optional[str]   = None
    category:         Optional[str]   = None
    description:      Optional[str]   = None
    quantity_on_hand: Optional[int]   = 0
    reorder_level:    Optional[int]   = 5
    unit_cost:        Optional[float] = None
    supplier:         Optional[str]   = None
    location:         Optional[str]   = None
    barcode:          Optional[str]   = None
    used_on_asset:    Optional[str]   = None
    unit:             Optional[str]   = "pcs"


def _dict(p: SparePart, show_cost: bool = True) -> dict:
    qty   = p.quantity_on_hand or 0
    reord = p.reorder_level    or 0
    d = {
        "id":               str(p.id),
        "part_code":        p.part_code,
        "name":             p.name,
        "category":         p.category,
        "description":      p.description,
        "quantity_on_hand": qty,
        "reorder_level":    reord,
        "supplier":         p.supplier,
        "location":         p.location,
        "barcode":          getattr(p, "barcode",        None),
        "used_on_asset":    getattr(p, "used_on_asset",  None),
        "photo_url":        getattr(p, "photo_url",      None),
        "has_photo":        bool(getattr(p, "photo_url", None)),
        "last_stock_take_at": p.last_stock_take_at.isoformat() if getattr(p, "last_stock_take_at", None) else None,
        "last_stock_take_by": getattr(p, "last_stock_take_by", None),
        "unit":             p.unit or "pcs",
        "status":           "Topup" if qty <= reord else "Enough",
        "is_low_stock":     qty <= reord,
        "created_at":       p.created_at.isoformat() if p.created_at else None,
        "updated_at":       p.updated_at.isoformat() if p.updated_at else None,
    }
    if show_cost:
        d["unit_cost"]    = p.unit_cost
        d["total_amount"] = round(qty * p.unit_cost, 2) if p.unit_cost else None
    return d


def _can_view_cost(user: User) -> bool:
    return user.role.value in ("admin", "manager")


def _require_admin_or_manager(user: User):
    if user.role.value not in ("admin", "manager"):
        raise HTTPException(403, "Only admins and managers can manage inventory items")


async def _get(part_id: UUID, db: AsyncSession) -> SparePart:
    p = (await db.execute(
        select(SparePart).where(SparePart.id == part_id)
    )).scalar_one_or_none()
    if not p:
        raise HTTPException(404, "Part not found")
    return p


def _safe_set(obj, field: str, value):
    """Set attribute only if the column exists on the model."""
    try:
        setattr(obj, field, value)
    except AttributeError:
        pass


# ══════════════════════════════════════════════════════════════════════════
# STATIC ROUTES — must be declared BEFORE /{part_id}
# ══════════════════════════════════════════════════════════════════════════

@router.get("/import/template")
async def download_template(_: User = Depends(get_current_user)):
    """Generate and return the inventory Excel import template."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(500, "openpyxl not installed — run: pip install openpyxl")

    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory"

    thin     = Side(style="thin", color="CCCCCC")
    border   = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", start_color="1E3A5F")
    opt_fill = PatternFill("solid", start_color="2D5986")
    req_hint = PatternFill("solid", start_color="EBF3FB")
    opt_hint = PatternFill("solid", start_color="F0FDF4")
    note_fill= PatternFill("solid", start_color="FFF9C4")
    data_font= Font(name="Arial", size=10)
    hdr_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    hint_font= Font(name="Arial", italic=True, color="5A6A7A", size=9)

    # Row 1 title
    ws.merge_cells("A1:L1")
    ws["A1"] = "CMMS Pro — Inventory Import Template"
    ws["A1"].font      = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    ws["A1"].fill      = hdr_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Row 2 note
    ws.merge_cells("A2:L2")
    ws["A2"] = (
        "Required columns (*): category, No, Part Item, Quantity, Threshold.  "
        "Leave 'Total Amount (RM)' and 'Status' blank — system calculates automatically."
    )
    ws["A2"].font      = Font(name="Arial", italic=True, size=10, color="5A6A7A")
    ws["A2"].fill      = note_fill
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 6

    # Headers row 4 — YOUR exact column names
    cols = [
        ("category *",           True,  "Electrical / Mechanical / Pneumatic / Hydraulic / Consumable / Other"),
        ("No *",                 True,  "Unique part number e.g. E-001, M-002  (must not duplicate)"),
        ("Part Item *",          True,  "Full part name e.g. Bearing 6205, Fuse 10A"),
        ("Description",          False, "Additional spec or model number (optional)"),
        ("Quantity *",           True,  "Current stock on hand — number only"),
        ("Rack No",              False, "Storage location e.g. Rack A1, Shelf B3"),
        ("Use at",               False, "Machine or asset name this part is used on"),
        ("Unit Price (RM)",      False, "Cost per unit in RM e.g. 12.50"),
        ("Total Amount(RM)",     False, "AUTO-CALCULATED — leave blank"),
        ("Barcode",              False, "Barcode or QR number (optional)"),
        ("Threshold *",          True,  "Min stock level — alert when Quantity ≤ Threshold"),
        ("Status(Enough/Topup)", False, "AUTO-CALCULATED — leave blank"),
    ]
    for c, (h, req, hint) in enumerate(cols, 1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.font      = hdr_font
        cell.fill      = hdr_fill if req else opt_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = border
        hc = ws.cell(row=5, column=c, value=hint)
        hc.font      = hint_font
        hc.fill      = req_hint if req else opt_hint
        hc.alignment = Alignment(wrap_text=True, vertical="top")
        hc.border    = border

    ws.row_dimensions[4].height = 24
    ws.row_dimensions[5].height = 40

    # Sample rows
    samples = [
        ["Electrical","E-001","Fuse 10A",           "Glass fuse 10A 250V",   50,"Rack A1","All machines",    0.80,"","BC001001",10,""],
        ["Electrical","E-002","Fuse 16A",           "Glass fuse 16A 250V",   30,"Rack A1","Compressor",       1.20,"","BC001002",10,""],
        ["Electrical","E-003","Contactor 9A",       "LC1D09 24VDC coil",      5,"Rack A2","Conveyor Motor",  45.00,"","BC001003", 2,""],
        ["Mechanical","M-001","Bearing 6205",       "Deep groove 25x52mm",   20,"Shelf B1","Pump A",           8.50,"","BC002001", 5,""],
        ["Mechanical","M-002","Bearing 6206",       "Deep groove 30x62mm",   15,"Shelf B1","Motor #1",         9.20,"","BC002002", 5,""],
        ["Mechanical","M-003","V-Belt A42",         "Classical V-belt",        8,"Shelf B2","Compressor",     12.00,"","BC002003", 3,""],
        ["Pneumatic", "P-001","SMC Filter Element", "AF20-F02 element",        6,"Rack C1","Pneumatic line",  22.00,"","BC003001", 2,""],
        ["Pneumatic", "P-002","Solenoid Valve 5/2", "SY3120-5LZD-M5",          4,"Rack C2","Cylinder line",  85.00,"","BC003002", 2,""],
        ["Consumable","C-001","Cable Tie 300mm",    "Black nylon 300mm",     200,"Rack D1","General",          0.05,"","BC004001",50,""],
        ["Consumable","C-002","WD-40 Spray 400ml",  "Multi-use lubricant",    10,"Rack D2","General",         12.00,"","BC004002", 3,""],
    ]
    alt = [PatternFill("solid",start_color="FFFFFF"), PatternFill("solid",start_color="F7FBFF")]
    for r, row in enumerate(samples, 6):
        for c, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=v if v != "" else None)
            cell.font = data_font; cell.fill = alt[r%2]; cell.border = border
            cell.alignment = Alignment(vertical="center", horizontal="right" if c in (5,8,9,11) else "left")
        ws.row_dimensions[r].height = 18
    for r in range(16, 56):
        for c in range(1, 13):
            ws.cell(row=r,column=c).fill = alt[r%2]
            ws.cell(row=r,column=c).border = border
        ws.row_dimensions[r].height = 18

    for i, w in enumerate([16,10,28,32,10,14,24,14,16,16,12,18], 1):
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A6"

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="CMMS_Inventory_Import_Template.xlsx"'},
    )


@router.post("/import/excel")
async def import_excel(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Import inventory from Excel. Accepts your existing format."""
    _require_admin_or_manager(current_user)
    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(400, "Only .xlsx files accepted")
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(500, "openpyxl not installed")

    content = await file.read()
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(400, "Cannot open file — must be a valid .xlsx file")

    ws = wb.worksheets[0]

    # Find header row (contains "No" or "Part Item")
    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=10):
        vals = [str(c.value or "").strip().lower() for c in row]
        if any(v in ("no","no *","part item","part item *","part_code") for v in vals):
            header_row = row[0].row
            break
    if not header_row:
        raise HTTPException(400,
            "Cannot find header row. Make sure row 4 has columns: "
            "No, Part Item, Quantity, Threshold.")

    # Column alias map
    aliases = {
        "category":"category","category *":"category",
        "no":"part_code","no *":"part_code","part_code":"part_code",
        "part item":"name","part item *":"name","name":"name",
        "description":"description",
        "quantity":"quantity","quantity *":"quantity","qty":"quantity",
        "rack no":"location","rack":"location","location":"location",
        "use at":"used_on","used at":"used_on",
        "unit price (rm)":"unit_cost","unit price":"unit_cost","price (rm)":"unit_cost",
        "total amount(rm)":"_skip","total amount (rm)":"_skip","total":"_skip",
        "barcode":"barcode",
        "threshold":"reorder_level","thereshold":"reorder_level",
        "threshold *":"reorder_level","min stock":"reorder_level",
        "status(enough/topup)":"_skip","status":"_skip",
        "supplier":"supplier","unit":"unit",
    }
    col_map = {}
    for cell in ws[header_row]:
        if cell.value:
            key = str(cell.value).strip().lower()
            mapped = aliases.get(key)
            if mapped and mapped != "_skip":
                col_map[mapped] = cell.column - 1

    missing = [r for r in ["part_code","name","quantity"] if r not in col_map]
    if missing:
        raise HTTPException(400,
            f"Missing columns: {', '.join(missing)}. "
            f"Found: {list(col_map.keys())}")

    existing = {p.part_code: p for p in (await db.execute(select(SparePart))).scalars().all()}
    # Default categories — admin/manager-managed, shown first in dropdowns.
    # Any category found in the Excel file that ISN'T in this list is not
    # discarded or forced into "Other" — it's kept exactly as written and
    # simply becomes a new category going forward (see matching logic below).
    # Default categories — admin/manager-managed, shown first in dropdowns.
    # Any category found in the Excel file that ISN'T in this list is not
    # discarded or forced into "Other" — it's kept exactly as written and
    # simply becomes a new category going forward (see matching logic below).
    valid_cats = set(DEFAULT_CATEGORIES)

    inserted, updated, skipped = 0, 0, []
    seen_codes_this_import: dict[str, int] = {}  # base_code -> count seen so far, to detect collisions
    disambiguated = 0  # rows whose "No" collided with another row and got a -2/-3... suffix

    def gcol(rv, key, default=""):
        idx = col_map.get(key)
        if idx is None or idx >= len(rv): return default
        v = rv[idx]; return str(v).strip() if v is not None else default

    def gnum(rv, key, default=None):
        raw = gcol(rv, key, "")
        try: return float(raw) if raw else default
        except: return default

    for rv in ws.iter_rows(min_row=header_row+1, values_only=True):
        if all(v is None or str(v).strip()=='' for v in rv): continue

        raw_no = gcol(rv,"part_code")
        name   = gcol(rv,"name")
        if not raw_no:
            skipped.append({"reason":"Missing part No"}); continue
        if not name:
            skipped.append({"part_code":raw_no,"reason":"Missing Part Item name"}); continue

        cat_raw  = gcol(rv,"category") or "Other"
        category = cat_raw.strip()  # keep the imported category as-is by default
        for vc in valid_cats:
            # If it's a close/fuzzy match to one of the default categories,
            # normalize it to that exact default spelling (so "bearings"
            # and "Bearing" don't become two different categories).
            if vc.lower()==category.lower() or vc.lower() in category.lower():
                category=vc; break
        # Anything that doesn't match a default category is kept exactly as
        # typed in the spreadsheet — it simply becomes a new category that
        # will show up everywhere (filters, dropdowns) going forward.

        # IMPORTANT: the "No" column commonly restarts at 1 for every category
        # section in real-world spreadsheets (Mechanical 1,2,3... Pneumatic
        # 1,2,3... etc). Using "No" alone as the unique part_code causes
        # later categories to silently overwrite earlier ones with the same
        # number. Combine category + No into the stored part_code so each
        # row stays distinct, while the prefix code abbreviation keeps it
        # short. If the sheet's "No" values are already globally unique
        # (e.g. "E-001", "M-002"), this just adds a short category prefix
        # and remains unique.
        # Generate a short, readable prefix for ANY category, not just a
        # fixed predefined set — so brand-new categories from an Excel
        # import still get a sensible part_code prefix instead of always
        # falling back to "OTH". Uses first 3 letters of the first word,
        # plus the first letter of each subsequent word — e.g.
        # "Bearing" -> "BEA", "Fastener and Hardware" -> "FASAH"
        # (kept short but distinct, reducing collisions between similarly
        # named categories like "Electrical/Electronic Components").
        _words = [w for w in category.split() if w]
        if _words:
            cat_prefix = (_words[0][:3] + "".join(w[0] for w in _words[1:])).upper()
        else:
            cat_prefix = "OTH"
        base_code = f"{cat_prefix}-{raw_no}"

        # Spreadsheets commonly reuse the same "No" across multiple sub-groups
        # within one category (e.g. Mechanical bearings 1,2,3... then
        # Mechanical belts 1,2,3... again). Detect collisions within THIS
        # import and disambiguate with a -2, -3... suffix so every row still
        # becomes its own distinct part instead of silently overwriting the
        # previous row with the same code.
        seen_count = seen_codes_this_import.get(base_code, 0)
        if seen_count == 0 and base_code not in existing:
            part_code = base_code
        else:
            # Either collided within this import, or already exists from a
            # previous import — count up until we find a free suffix.
            n = max(seen_count, 1)
            part_code = f"{base_code}-{n+1}"
            while part_code in existing:
                n += 1
                part_code = f"{base_code}-{n+1}"
            disambiguated += 1
        seen_codes_this_import[base_code] = seen_count + 1

        qty           = int(gnum(rv,"quantity",0) or 0)
        uc_raw        = gnum(rv,"unit_cost")
        unit_cost     = round(float(uc_raw),4) if uc_raw is not None else None
        reorder_level = int(gnum(rv,"reorder_level",5) or 5)
        description   = gcol(rv,"description")  or None
        location      = gcol(rv,"location")     or None
        used_on       = gcol(rv,"used_on")      or None
        barcode       = gcol(rv,"barcode")      or None
        supplier      = gcol(rv,"supplier")     or None
        unit          = gcol(rv,"unit")         or "pcs"

        if part_code in existing:
            p = existing[part_code]
            p.name             = name
            p.category         = category
            p.description      = description
            p.quantity_on_hand = qty
            p.reorder_level    = reorder_level
            if unit_cost  is not None: p.unit_cost = unit_cost
            if location:  p.location  = location
            if supplier:  p.supplier  = supplier
            p.unit = unit
            _safe_set(p,"barcode",       barcode)
            _safe_set(p,"used_on_asset", used_on)
            p.updated_at = datetime.utcnow()
            updated += 1
        else:
            p = SparePart(
                part_code        = part_code,
                name             = name,
                category         = category,
                description      = description,
                quantity_on_hand = qty,
                reorder_level    = reorder_level,
                unit_cost        = unit_cost,
                location         = location,
                supplier         = supplier,
                unit             = unit,
            )
            _safe_set(p,"barcode",       barcode)
            _safe_set(p,"used_on_asset", used_on)
            db.add(p)
            existing[part_code] = p
            inserted += 1

    await db.flush()
    return {
        "success":True,"inserted":inserted,"updated":updated,
        "skipped":len(skipped),"skipped_details":skipped,
        "message":f"Import complete: {inserted} added, {updated} updated, {len(skipped)} skipped.",
    }


@router.get("/export/excel")
async def export_excel(
    category:  Optional[str] = None,
    low_stock: bool = False,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Export inventory — one sheet per category, plus a Summary sheet totaling all categories."""
    _require_admin_or_manager(current_user)
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(500,"openpyxl not installed")

    q = select(SparePart).order_by(SparePart.category, SparePart.name)
    if category:  q = q.where(SparePart.category == category)
    if low_stock: q = q.where(SparePart.quantity_on_hand <= SparePart.reorder_level)
    parts = (await db.execute(q)).scalars().all()

    thin     = Side(style="thin",color="CCCCCC")
    border   = Border(left=thin,right=thin,top=thin,bottom=thin)
    hdr_font = Font(name="Arial",bold=True,color="FFFFFF",size=10)
    hdr_fill = PatternFill("solid",start_color="1E3A5F")
    title_font = Font(name="Arial",bold=True,size=11,color="FFFFFF")
    data_font= Font(name="Arial",size=10)
    bold_font= Font(name="Arial",bold=True,size=10)
    enough   = PatternFill("solid",start_color="DCFCE7")
    topup    = PatternFill("solid",start_color="FEF9C4")
    summary_fill = PatternFill("solid",start_color="EFF6FF")

    # Group parts by category, preserving a stable, sorted category order
    by_category: dict[str, list] = {}
    for p in parts:
        cat = p.category or "Uncategorized"
        by_category.setdefault(cat, []).append(p)
    sorted_categories = sorted(by_category.keys())

    def _sheet_name(cat: str) -> str:
        # Excel sheet names: max 31 chars, no \/?*[]:
        name = re.sub(r'[\\/?*\[\]:]', '-', cat)[:31]
        return name or "Uncategorized"

    wb = Workbook()
    wb.remove(wb.active)  # we'll add Summary first, then one sheet per category

    headers = ["No","Part Item","Description","Quantity",
               "Rack No","Use at","Unit Price (RM)","Total Amount(RM)",
               "Barcode","Threshold","Status(Enough/Topup)"]
    col_widths = [16,28,32,10,14,24,14,16,16,12,18]

    sheet_refs: list[tuple[str, str, int]] = []  # (category, sheet_name, last_data_row) for the Summary sheet

    for cat in sorted_categories:
        cat_parts = sorted(by_category[cat], key=lambda x: x.name)
        sheet_name = _sheet_name(cat)
        # Guard against duplicate sheet names after sanitization
        base_name, n = sheet_name, 2
        while sheet_name in wb.sheetnames:
            sheet_name = f"{base_name[:28]}-{n}"; n += 1

        ws = wb.create_sheet(sheet_name)

        ws.merge_cells("A1:K1")
        ws["A1"] = f"{cat}  |  CMMS Pro Inventory  |  {datetime.utcnow().strftime('%d %b %Y %H:%M')} UTC  |  {len(cat_parts)} parts"
        ws["A1"].font = title_font
        ws["A1"].fill = hdr_fill
        ws["A1"].alignment = Alignment(horizontal="center")
        ws.row_dimensions[1].height = 26

        for c,h in enumerate(headers,1):
            cell = ws.cell(row=2,column=c,value=h)
            cell.font=hdr_font; cell.fill=hdr_fill; cell.border=border
            cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
        ws.row_dimensions[2].height = 24

        topup_n=enough_n=0
        first_data_row = 3
        for r,p in enumerate(cat_parts,first_data_row):
            qty   = p.quantity_on_hand or 0
            reord = p.reorder_level    or 0
            sts   = "Topup" if qty<=reord else "Enough"
            if sts=="Topup": topup_n+=1
            else: enough_n+=1
            row_data = [
                p.part_code, p.name,
                p.description or "",
                qty,
                p.location or "",
                getattr(p,"used_on_asset",None) or "",
                p.unit_cost if p.unit_cost else "",
                None,  # Total Amount(RM) — filled as a formula below
                getattr(p,"barcode",None) or "",
                reord, sts,
            ]
            fill = enough if sts=="Enough" else topup
            for c,v in enumerate(row_data,1):
                cell=ws.cell(row=r,column=c,value=v)
                cell.font=data_font; cell.fill=fill; cell.border=border
                cell.alignment=Alignment(vertical="center",horizontal="right" if c in(4,7,8,10) else "left")
            # Total Amount(RM) = Quantity * Unit Price, as a live formula
            ws.cell(row=r,column=8,value=f"=D{r}*G{r}").font=data_font
            ws.cell(row=r,column=8).fill=fill
            ws.cell(row=r,column=8).border=border
            ws.cell(row=r,column=8).alignment=Alignment(vertical="center",horizontal="right")
            ws.row_dimensions[r].height=18

        last_data_row = first_data_row + len(cat_parts) - 1 if cat_parts else first_data_row - 1
        foot = last_data_row + 1
        ws.cell(row=foot,column=1,value="TOTAL").font=bold_font
        if cat_parts:
            ws.cell(row=foot,column=4,value=f"=SUM(D{first_data_row}:D{last_data_row})").font=bold_font
            ws.cell(row=foot,column=8,value=f"=SUM(H{first_data_row}:H{last_data_row})").font=bold_font
        else:
            ws.cell(row=foot,column=4,value=0).font=bold_font
            ws.cell(row=foot,column=8,value=0).font=bold_font
        ws.cell(row=foot,column=11,value=f"Enough: {enough_n}   Topup: {topup_n}").font=Font(name="Arial",bold=True,color="1E3A5F")
        ws.row_dimensions[foot].height=20

        for i,w in enumerate(col_widths,1):
            ws.column_dimensions[get_column_letter(i)].width=w
        ws.freeze_panes="A3"

        sheet_refs.append((cat, sheet_name, foot))

    # ── Summary sheet — totals every category, placed first ──
    summary = wb.create_sheet("Summary", 0)
    summary.merge_cells("A1:D1")
    summary["A1"] = f"CMMS Pro — Inventory Summary  |  {datetime.utcnow().strftime('%d %b %Y %H:%M')} UTC  |  {len(sorted_categories)} categories  |  {len(parts)} parts total"
    summary["A1"].font = title_font
    summary["A1"].fill = hdr_fill
    summary["A1"].alignment = Alignment(horizontal="center")
    summary.row_dimensions[1].height = 26

    summary_headers = ["Category","Total Quantity","Total Inventory Cost (RM)","Part Count"]
    for c,h in enumerate(summary_headers,1):
        cell = summary.cell(row=2,column=c,value=h)
        cell.font=hdr_font; cell.fill=hdr_fill; cell.border=border
        cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    summary.row_dimensions[2].height = 24

    first_summary_row = 3
    for i,(cat, sheet_name, foot_row) in enumerate(sheet_refs):
        r = first_summary_row + i
        summary.cell(row=r,column=1,value=cat).font=data_font
        # Pull totals live from each category sheet via cross-sheet formula references
        summary.cell(row=r,column=2,value=f"='{sheet_name}'!D{foot_row}").font=data_font
        summary.cell(row=r,column=3,value=f"='{sheet_name}'!H{foot_row}").font=data_font
        summary.cell(row=r,column=4,value=len(by_category[cat])).font=data_font
        for c in (1,2,3,4):
            cell=summary.cell(row=r,column=c)
            cell.fill=summary_fill; cell.border=border
            cell.alignment=Alignment(vertical="center",horizontal="right" if c in(2,3,4) else "left")
        summary.row_dimensions[r].height=18

    last_summary_row = first_summary_row + len(sheet_refs) - 1 if sheet_refs else first_summary_row - 1
    grand_foot = last_summary_row + 1
    summary.cell(row=grand_foot,column=1,value="GRAND TOTAL").font=bold_font
    if sheet_refs:
        summary.cell(row=grand_foot,column=2,value=f"=SUM(B{first_summary_row}:B{last_summary_row})").font=bold_font
        summary.cell(row=grand_foot,column=3,value=f"=SUM(C{first_summary_row}:C{last_summary_row})").font=bold_font
        summary.cell(row=grand_foot,column=4,value=f"=SUM(D{first_summary_row}:D{last_summary_row})").font=bold_font
    else:
        for c in (2,3,4):
            summary.cell(row=grand_foot,column=c,value=0).font=bold_font
    for c in (1,2,3,4):
        summary.cell(row=grand_foot,column=c).fill=hdr_fill
        summary.cell(row=grand_foot,column=c).font=Font(name="Arial",bold=True,color="FFFFFF")
    summary.row_dimensions[grand_foot].height=20

    for i,w in enumerate([28,18,26,14],1):
        summary.column_dimensions[get_column_letter(i)].width=w
    summary.freeze_panes="A3"

    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    fname=f"Inventory_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition":f'attachment; filename="{fname}"'})


# ══════════════════════════════════════════════════════════════════════════
# PATH PARAMETER ROUTES — after all static routes
# ══════════════════════════════════════════════════════════════════════════

@router.get("/categories")
async def list_categories(
    db: AsyncSession = Depends(get_db), _: User = Depends(get_current_user),
):
    """
    Always includes the default category list (so they're selectable even
    before any part uses them yet), plus any additional categories that
    already exist in the data — e.g. ones introduced by an Excel import
    that didn't match a default category. Default categories are listed
    first, in their defined order; anything extra is appended after,
    alphabetically.
    """
    result = await db.execute(
        select(SparePart.category)
        .where(SparePart.category.isnot(None), SparePart.category != "")
        .distinct()
    )
    existing = {row[0] for row in result.all()}
    extra = sorted(existing - set(DEFAULT_CATEGORIES))
    return DEFAULT_CATEGORIES + extra


@router.get("/stock-take-status")
async def stock_take_status(
    stale_after_days: int = 30,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Reports which parts have never been physically verified via the Stock
    Take tab, and which were checked but are now "stale" (older than
    stale_after_days). Sorted so the most overdue / never-checked items
    appear first — gives a direct answer to "what have I missed".
    """
    _require_admin_or_manager(current_user)
    result = await db.execute(select(SparePart).order_by(SparePart.category, SparePart.name))
    parts = result.scalars().all()

    now = datetime.utcnow()
    cutoff = now - timedelta(days=stale_after_days)

    never_checked, stale, up_to_date = [], [], []
    for p in parts:
        last = getattr(p, "last_stock_take_at", None)
        entry = {
            "id": str(p.id), "part_code": p.part_code, "name": p.name,
            "category": p.category, "location": p.location,
            "last_stock_take_at": last.isoformat() if last else None,
            "last_stock_take_by": getattr(p, "last_stock_take_by", None),
            "days_since": (now - last).days if last else None,
        }
        if not last:
            never_checked.append(entry)
        elif last < cutoff:
            stale.append(entry)
        else:
            up_to_date.append(entry)

    # Most overdue first within each bucket
    stale.sort(key=lambda x: x["days_since"], reverse=True)
    up_to_date.sort(key=lambda x: x["days_since"], reverse=True)

    return {
        "stale_after_days": stale_after_days,
        "total_parts": len(parts),
        "never_checked_count": len(never_checked),
        "stale_count": len(stale),
        "up_to_date_count": len(up_to_date),
        "never_checked": never_checked,
        "stale": stale,
        "up_to_date": up_to_date,
    }


@router.get("/")
async def list_parts(
    low_stock:bool=False, category:Optional[str]=None,
    skip:int=0, limit:int=500,
    db:AsyncSession=Depends(get_db), current_user:User=Depends(get_current_user),
):
    q = select(SparePart)
    if low_stock: q=q.where(SparePart.quantity_on_hand<=SparePart.reorder_level)
    if category:  q=q.where(SparePart.category==category)
    q=q.offset(skip).limit(limit).order_by(SparePart.category,SparePart.name)
    show_cost=_can_view_cost(current_user)
    return [_dict(p,show_cost) for p in (await db.execute(q)).scalars().all()]


@router.post("/", status_code=201)
async def create_part(
    body:PartIn, db:AsyncSession=Depends(get_db), current_user:User=Depends(get_current_user),
):
    _require_admin_or_manager(current_user)
    if not body.part_code: raise HTTPException(400,"part_code required")
    if not body.name:      raise HTTPException(400,"name required")
    if (await db.execute(select(SparePart).where(SparePart.part_code==body.part_code))).scalar_one_or_none():
        raise HTTPException(400,f"Part code '{body.part_code}' already exists")
    p=SparePart(
        part_code=body.part_code, name=body.name,
        category=body.category or "Other",
        description=body.description,
        quantity_on_hand=body.quantity_on_hand or 0,
        reorder_level=body.reorder_level or 5,
        unit_cost=body.unit_cost, supplier=body.supplier,
        location=body.location, unit=body.unit or "pcs",
    )
    _safe_set(p,"barcode",      body.barcode)
    _safe_set(p,"used_on_asset",body.used_on_asset)
    db.add(p); await db.flush(); await db.refresh(p)
    await ws_manager.broadcast_event(ROOM,"inventory.created",{"id":str(p.id),"name":p.name})
    return _dict(p,_can_view_cost(current_user))


@router.get("/{part_id}")
async def get_part(part_id:UUID,db:AsyncSession=Depends(get_db),current_user:User=Depends(get_current_user)):
    return _dict(await _get(part_id,db),_can_view_cost(current_user))


@router.patch("/{part_id}")
async def update_part(
    part_id:UUID, body:PartIn,
    db:AsyncSession=Depends(get_db), current_user:User=Depends(get_current_user),
):
    _require_admin_or_manager(current_user)
    p=await _get(part_id,db)
    updates=body.model_dump(exclude_unset=True,exclude_none=True)
    for k,v in updates.items():
        if k in("barcode","used_on_asset"): _safe_set(p,k,v)
        else:
            try: setattr(p,k,v)
            except: pass
    p.updated_at=datetime.utcnow()
    await db.flush()
    await ws_manager.broadcast_event(ROOM,"inventory.updated",{"id":str(p.id),"name":p.name})
    return _dict(p,_can_view_cost(current_user))


@router.delete("/{part_id}",status_code=204)
async def delete_part(part_id:UUID,db:AsyncSession=Depends(get_db),current_user:User=Depends(get_current_user)):
    _require_admin_or_manager(current_user)
    p=await _get(part_id,db)
    await db.delete(p)
    await ws_manager.broadcast_event(ROOM,"inventory.deleted",{"id":str(part_id)})


@router.post("/{part_id}/restock")
async def restock(
    part_id:UUID, quantity:int,
    db:AsyncSession=Depends(get_db), current_user:User=Depends(get_current_user),
):
    if quantity<1: raise HTTPException(400,"Quantity must be at least 1")
    p=await _get(part_id,db)
    p.quantity_on_hand+=quantity
    p.updated_at=datetime.utcnow()
    await db.flush()
    await ws_manager.broadcast_event(ROOM,"inventory.restocked",{"id":str(p.id),"name":p.name,"new_qty":p.quantity_on_hand})
    return _dict(p,_can_view_cost(current_user))


class PhotoUpload(BaseModel):
    data: str  # base64 data URL, e.g. "data:image/jpeg;base64,/9j/4AAQ..."


@router.put("/{part_id}/photo")
async def upload_part_photo(
    part_id: UUID, body: PhotoUpload,
    db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user),
):
    """
    Upload or replace the reference photo for a part. Stored as a base64
    data URL directly on the row (same lightweight approach used for repair
    request photos) — no filesystem dependency.
    Max ~800KB raw image to keep page loads reasonable for the parts table.
    """
    _require_admin_or_manager(current_user)
    p = await _get(part_id, db)

    data_url = body.data if body.data.startswith("data:") else f"data:image/jpeg;base64,{body.data}"
    raw_b64 = data_url.split(",", 1)[1] if "," in data_url else data_url
    size_kb = len(raw_b64) * 0.75 / 1024
    if size_kb > 800:
        raise HTTPException(400, f"Photo too large ({size_kb:.0f}KB) — please use an image under ~800KB")

    p.photo_url = data_url
    p.updated_at = datetime.utcnow()
    await db.flush()
    await ws_manager.broadcast_event(ROOM, "inventory.updated", {"id": str(p.id), "name": p.name})
    return _dict(p, _can_view_cost(current_user))


@router.delete("/{part_id}/photo")
async def delete_part_photo(
    part_id: UUID,
    db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user),
):
    _require_admin_or_manager(current_user)
    p = await _get(part_id, db)
    p.photo_url = None
    p.updated_at = datetime.utcnow()
    await db.flush()
    await ws_manager.broadcast_event(ROOM, "inventory.updated", {"id": str(p.id), "name": p.name})
    return _dict(p, _can_view_cost(current_user))