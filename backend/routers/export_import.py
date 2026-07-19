"""
Export & Import Router
=======================
All routes under a single /data prefix to avoid FastAPI route conflicts.

GET  /data/export/history.xlsx     — download work order history as Excel
GET  /data/import/template         — download the asset import template
POST /data/import/assets           — upload Excel to bulk create/update assets
"""

import io
import os
import re
import base64
from datetime import datetime
from typing import Optional
from uuid import UUID

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse, FileResponse
from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from db import get_db
from models import Asset, WorkOrder, WorkOrderStatus, WorkOrderType, Priority, AssetStatus, User, TaskAssignment
from auth import get_current_user, forbid_viewer
from routers.analysis import MY_TZ_OFFSET

# Single router — no prefix conflict
router = APIRouter(prefix="/data", tags=["export_import"], dependencies=[Depends(forbid_viewer)])


# ─────────────────────────────────────────────────────────────────────────────
# TEMPLATE DOWNLOAD  GET /data/import/template
# Must be declared BEFORE the POST /data/import/assets route
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/import/template")
async def download_asset_template(
    _: User = Depends(get_current_user),
):
    """Serve the asset import Excel template."""
    # Look for a pre-built template file first
    base = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    candidates = [
        os.path.join(base, "CMMS_Asset_Import_Template.xlsx"),
        os.path.join(base, "..", "CMMS_Asset_Import_Template.xlsx"),
        "/opt/cmms/CMMS_Asset_Import_Template.xlsx",
    ]
    for path in candidates:
        if os.path.exists(path):
            return FileResponse(
                path,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                filename="CMMS_Asset_Import_Template.xlsx",
            )

    # Generate template on the fly if file not found
    return _generate_template()


def _generate_template():
    """Build and return the Excel template as a streaming response."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(500, "openpyxl not installed. Run: pip install openpyxl")

    wb  = Workbook()
    ws  = wb.active
    ws.title = "Asset Import"

    thin   = Side(style="thin",   color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Row 1: Title
    ws.merge_cells("A1:J1")
    ws["A1"] = "CMMS Pro — Asset Import Template"
    ws["A1"].font      = Font(name="Arial", bold=True, size=13, color="1E3A5F")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].fill      = PatternFill("solid", start_color="EBF3FB")
    ws.row_dimensions[1].height = 30

    # Row 2: Instructions
    ws.merge_cells("A2:J2")
    ws["A2"] = (
        "Required columns: asset_code, name, category, location.  "
        "Do NOT modify column headers.  "
        "Save as .xlsx before uploading."
    )
    ws["A2"].font      = Font(name="Arial", italic=True, size=10, color="5A6A7A")
    ws["A2"].alignment = Alignment(horizontal="center")
    ws["A2"].fill      = PatternFill("solid", start_color="FFF9C4")
    ws.row_dimensions[2].height = 18

    # Row 3: spacer
    ws.row_dimensions[3].height = 6

    # Row 4: Headers
    headers = [
        "asset_code *", "name *", "category *", "location *",
        "manufacturer", "model", "serial_number",
        "status", "purchase_date", "notes",
    ]
    hints = [
        "Unique code e.g. PMP-001",
        "Full machine name",
        "Multi Handling System (MHS) / Machine / Robot / Forklift (Gas) / Forklift (Battery) / Forklift (Diesel) / Utilities / Other",
        "Physical location e.g. Building A - Ground Floor",
        "Brand name (optional)",
        "Model number (optional)",
        "Serial or tag number (optional)",
        "operational (default) / under_maintenance / out_of_service / decommissioned",
        "YYYY-MM-DD  (optional)",
        "Any additional notes (optional)",
    ]
    hdr_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    hdr_fill = PatternFill("solid", start_color="1E3A5F")
    hint_fill = PatternFill("solid", start_color="EBF3FB")

    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.font = hdr_font; cell.fill = hdr_fill; cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        hcell = ws.cell(row=5, column=c, value=hints[c-1])
        hcell.font = Font(name="Arial", italic=True, color="5A6A7A", size=9)
        hcell.fill = hint_fill; hcell.border = border
        hcell.alignment = Alignment(wrap_text=True, vertical="top")

    ws.row_dimensions[4].height = 26
    ws.row_dimensions[5].height = 34

    # Sample data rows 6-10
    samples = [
        ["PMP-001","Centrifugal Pump A","Pump","Building A - Ground Floor","Grundfos","CM10-4","GF-2021-00341","operational","2021-03-15","Primary water supply pump"],
        ["MTR-001","Conveyor Drive Motor #1","Motor","Production Line 1","ABB","M2BAX 90","ABB-MTR-9921","operational","2020-08-10",""],
        ["CMP-001","Air Compressor Unit 1","Compressor","Utility Room B","Atlas Copco","GA15","AC-GA15-4423","operational","2019-06-01",""],
        ["HVAC-001","AHU - Production Hall","HVAC","Rooftop Level 2","Carrier","AHU-30XA","CAR-30XA-881","operational","",""],
        ["GEN-001","Standby Generator","Generator","Genset Room - Block C","Caterpillar","C15","CAT-C15-0042","operational","2018-01-20","500 kVA"],
    ]
    fills = [
        PatternFill("solid", start_color="FFFFFF"),
        PatternFill("solid", start_color="F7FBFF"),
    ]
    for r, row in enumerate(samples, 6):
        for c, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = Font(name="Arial", size=10)
            cell.fill = fills[r % 2]
            cell.border = border
        ws.row_dimensions[r].height = 18

    # 15 blank rows for user data
    for r in range(11, 26):
        for c in range(1, 11):
            cell = ws.cell(row=r, column=c, value="")
            cell.fill = fills[r % 2]
            cell.border = border
        ws.row_dimensions[r].height = 18

    # Column widths
    for i, w in enumerate([16,28,20,28,18,14,18,22,14,32], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A6"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="CMMS_Asset_Import_Template.xlsx"'},
    )


# ─────────────────────────────────────────────────────────────────────────────
# ASSET IMPORT  POST /data/import/assets
# ─────────────────────────────────────────────────────────────────────────────

@router.post("/import/assets")
async def import_assets(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """
    Upload a filled-in Excel template to bulk create/update assets.
    - New asset_code  → INSERT
    - Existing code   → UPDATE
    Returns a summary: inserted, updated, skipped.
    """
    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(400, "Only .xlsx files are accepted. Please save your file as Excel Workbook (.xlsx).")

    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(500, "openpyxl not installed on server. Run: pip install openpyxl")

    content = await file.read()
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(400, "Cannot open file — make sure it is a valid .xlsx Excel file.")

    # Find the asset sheet (first sheet containing "asset" or "import" in name, else sheet 1)
    sheet_name = wb.sheetnames[0]
    for name in wb.sheetnames:
        if "asset" in name.lower() or "import" in name.lower():
            sheet_name = name
            break
    ws = wb[sheet_name]

    # Find header row — look for "asset_code" in any of the first 10 rows
    header_row_idx = None
    for row in ws.iter_rows(min_row=1, max_row=10):
        for cell in row:
            val = str(cell.value or "").lower().replace("*","").strip()
            if val == "asset_code":
                header_row_idx = cell.row
                break
        if header_row_idx:
            break

    if not header_row_idx:
        raise HTTPException(
            400,
            "Cannot find 'asset_code' header. "
            "Make sure you are using the official CMMS template and have not renamed the columns."
        )

    # Map column names to 0-based indices
    col_map = {}
    for cell in ws[header_row_idx]:
        if cell.value:
            key = str(cell.value).lower().replace("*","").strip()
            col_map[key] = cell.column - 1

    missing = [r for r in ["asset_code","name","category","location"] if r not in col_map]
    if missing:
        raise HTTPException(400, f"Missing required columns: {', '.join(missing)}")

    # Load existing assets for upsert
    existing = {a.asset_code: a for a in (await db.execute(select(Asset))).scalars().all()}

    valid_status   = {s.value for s in AssetStatus}
    valid_category = {"Multi Handling System (MHS)","Machine","Robot","Forklift (Gas)","Forklift (Battery)","Forklift (Diesel)","Utilities","Other"}

    inserted, updated, skipped = 0, 0, []

    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        # Skip blank rows
        if all(v is None or str(v).strip() == "" for v in row):
            continue

        def gcol(key):
            idx = col_map.get(key)
            if idx is None or idx >= len(row):
                return ""
            return str(row[idx] or "").strip()

        asset_code    = gcol("asset_code")
        name          = gcol("name")
        category      = gcol("category")
        location      = gcol("location")
        manufacturer  = gcol("manufacturer") or None
        model         = gcol("model")         or None
        serial_number = gcol("serial_number") or None
        notes         = gcol("notes")         or None
        status_val    = gcol("status").lower() or "operational"
        pd_raw        = gcol("purchase_date")

        # Validate required fields
        if not asset_code:
            skipped.append({"reason": "Missing asset_code (blank row skipped)"}); continue
        if not name:
            skipped.append({"asset_code": asset_code, "reason": "Missing name"}); continue
        if not location:
            skipped.append({"asset_code": asset_code, "reason": "Missing location"}); continue

        # Normalise category
        if category not in valid_category:
            category = "Other"

        # Normalise status
        if status_val not in valid_status:
            status_val = "operational"

        # Parse purchase_date
        purchase_date = None
        if pd_raw:
            for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"):
                try:
                    purchase_date = datetime.strptime(pd_raw, fmt)
                    break
                except ValueError:
                    pass

        if asset_code in existing:
            # UPDATE existing asset
            a = existing[asset_code]
            a.name          = name
            a.category      = category
            a.location      = location
            a.manufacturer  = manufacturer
            a.model         = model
            a.serial_number = serial_number
            a.status        = AssetStatus(status_val)
            a.notes         = notes
            if purchase_date:
                a.purchase_date = purchase_date
            updated += 1
        else:
            # INSERT new asset
            a = Asset(
                asset_code    = asset_code,
                name          = name,
                category      = category,
                location      = location,
                manufacturer  = manufacturer,
                model         = model,
                serial_number = serial_number,
                status        = AssetStatus(status_val),
                purchase_date = purchase_date,
                notes         = notes,
            )
            db.add(a)
            existing[asset_code] = a
            inserted += 1

    await db.flush()

    return {
        "success":         True,
        "inserted":        inserted,
        "updated":         updated,
        "skipped":         len(skipped),
        "skipped_details": skipped,
        "message":         f"Import complete: {inserted} inserted, {updated} updated, {len(skipped)} skipped.",
    }


# ─────────────────────────────────────────────────────────────────────────────
# WORK ORDER HISTORY IMPORT  POST /data/import/history
# Accepts the SAME 16-column layout produced by GET /data/export/history.xlsx
# ─────────────────────────────────────────────────────────────────────────────

HISTORY_HEADERS = [
    "WO Number", "Machine", "Asset Code", "Location", "Problem Category", "Problem Description",
    "Type", "Priority", "Status", "Operator",
    "Root Cause Analysis", "Action Taken", "Completed By",
    "Request Time (MYT)", "Completed Time (MYT)",
    "Downtime / Actual Hours", "Downtime Type",
]


def _parse_myt_datetime(raw):
    """
    Parses a "Request Time (MYT)" / "Completed Time (MYT)" cell back into a
    naive UTC datetime for storage — the reverse of how export_history
    writes those columns ((wo.created_at + MY_TZ_OFFSET).strftime(...)).
    Accepts the export's own "07 Jul 2026 20:42" text, or a real datetime
    value if Excel auto-converted the cell when the file was edited/saved.
    Returns None for blank/unparseable cells.
    """
    if raw is None or raw == "":
        return None
    if isinstance(raw, datetime):
        my_dt = raw
    else:
        text = str(raw).strip()
        if not text or text == "—":
            return None
        my_dt = None
        for fmt in ("%d %b %Y %H:%M", "%d %B %Y %H:%M", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%d/%m/%Y %H:%M"):
            try:
                my_dt = datetime.strptime(text, fmt)
                break
            except ValueError:
                continue
        if my_dt is None:
            return None
    return my_dt - MY_TZ_OFFSET


@router.post("/import/history")
async def import_history(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Bulk-imports historical work orders from an .xlsx file laid out exactly
    like GET /data/export/history.xlsx — easiest way to build one is to
    export your current history first (even with zero/few rows, to get the
    header row), then fill in past records underneath it.

    Rules:
      - "WO Number" is used as the unique key. If it already exists in the
        system, that row is SKIPPED (never overwritten) so re-uploading the
        same file twice is always safe. Leave it blank to auto-generate one.
      - "Asset Code" (preferred) or "Machine" (name, case-insensitive
        fallback) must match an existing asset — rows that don't match are
        skipped and reported, since silently creating placeholder assets
        would pollute the Asset Registry. If multiple assets share the same
        Machine name, "Location" is used to tell them apart; if it still
        can't be narrowed to exactly one, the row is skipped with a note
        asking for Asset Code or Location.
      - "Type" / "Priority" / "Status" fall back to sensible defaults
        (corrective / medium / open) if blank or not a recognised value,
        rather than failing the whole row.
      - The free-text fields (Problem Category/Description, Root Cause,
        Action Taken, Operator, Completed By) are written into the
        description using the same "[OPERATOR REQUEST]" / "[COMPLETION —
        ...]" block format the rest of the app already uses, prefixed with
        an "[IMPORTED HISTORICAL RECORD]" marker — so these rows show up
        correctly everywhere (History Log, downtime reports, re-export)
        exactly like normally-created work orders.
    """
    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(400, "Only .xlsx files are accepted. Please save your file as Excel Workbook (.xlsx).")

    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(500, "openpyxl not installed on server. Run: pip install openpyxl")

    content = await file.read()
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(400, "Cannot open file — make sure it is a valid .xlsx Excel file.")

    sheet_name = wb.sheetnames[0]
    for name in wb.sheetnames:
        if "history" in name.lower() or "work order" in name.lower() or "import" in name.lower():
            sheet_name = name
            break
    ws = wb[sheet_name]

    # Find the header row — look for "WO Number" in the first 10 rows
    # (the export puts a title + subtitle above it, same as the asset template).
    header_row_idx = None
    for row in ws.iter_rows(min_row=1, max_row=10):
        for cell in row:
            if str(cell.value or "").strip().lower() == "wo number":
                header_row_idx = cell.row
                break
        if header_row_idx:
            break

    if not header_row_idx:
        raise HTTPException(
            400,
            "Cannot find a 'WO Number' header. Use the same layout as "
            "'Export History' — download that first if you need a starting template."
        )

    col_map = {}
    for cell in ws[header_row_idx]:
        if cell.value:
            col_map[str(cell.value).strip().lower()] = cell.column - 1

    required = ["wo number", "machine", "type", "priority", "status"]
    missing = [h for h in required if h not in col_map]
    if missing:
        raise HTTPException(400, f"Missing required columns: {', '.join(missing)}")

    # Lookups for resolving assets and de-duping WO numbers
    from collections import defaultdict
    assets = (await db.execute(select(Asset))).scalars().all()
    assets_by_code = {a.asset_code.strip().lower(): a for a in assets if a.asset_code}
    assets_by_name = defaultdict(list)
    for a in assets:
        if a.name:
            assets_by_name[a.name.strip().lower()].append(a)
    existing_wo_numbers = {
        wo_number for (wo_number,) in (await db.execute(select(WorkOrder.wo_number))).all()
    }
    next_seq = (await db.execute(select(func.count()).select_from(WorkOrder))).scalar() or 0

    valid_types    = {t.value for t in WorkOrderType}
    valid_priority = {p.value for p in Priority}
    valid_status   = {s.value for s in WorkOrderStatus}

    def gcol(row, key):
        idx = col_map.get(key)
        if idx is None or idx >= len(row):
            return ""
        v = row[idx]
        return "" if v is None else str(v).strip()

    inserted, skipped = 0, []

    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if all(v is None or str(v).strip() == "" for v in row):
            continue

        wo_number      = gcol(row, "wo number")
        machine        = gcol(row, "machine")
        asset_code     = gcol(row, "asset code")
        location       = gcol(row, "location")
        category       = gcol(row, "problem category")
        problem        = gcol(row, "problem description")
        type_raw       = gcol(row, "type").lower()
        priority_raw   = gcol(row, "priority").lower()
        status_raw     = gcol(row, "status").lower()
        operator       = gcol(row, "operator")
        root_cause     = gcol(row, "root cause analysis")
        action_taken   = gcol(row, "action taken")
        completed_by   = gcol(row, "completed by")
        request_time   = gcol(row, "request time (myt)")
        completed_time = gcol(row, "completed time (myt)")
        actual_hours   = gcol(row, "downtime / actual hours")
        downtime_type  = gcol(row, "downtime type").lower()

        if wo_number and wo_number in existing_wo_numbers:
            skipped.append({"wo_number": wo_number, "reason": "WO Number already exists — skipped"})
            continue

        # Resolve asset: asset code first (unambiguous), then machine name —
        # using Location to disambiguate when multiple assets share a name.
        asset = assets_by_code.get(asset_code.strip().lower()) if asset_code else None
        if not asset and machine:
            candidates = assets_by_name.get(machine.strip().lower(), [])
            if len(candidates) == 1:
                asset = candidates[0]
            elif len(candidates) > 1 and location:
                matches = [a for a in candidates if (a.location or "").strip().lower() == location.strip().lower()]
                if len(matches) == 1:
                    asset = matches[0]
        if not asset:
            candidates = assets_by_name.get(machine.strip().lower(), []) if machine else []
            reason = (
                f"Multiple assets named '{machine}' — add Asset Code or a matching Location to disambiguate"
                if len(candidates) > 1
                else f"No matching asset for Asset Code '{asset_code}' / Machine '{machine}'"
            )
            skipped.append({
                "wo_number": wo_number or "(blank)",
                "reason": reason,
            })
            continue

        wo_type  = type_raw     if type_raw     in valid_types    else "corrective"
        wo_pri   = priority_raw if priority_raw in valid_priority else "medium"
        wo_stat  = status_raw   if status_raw   in valid_status   else "open"

        created_dt   = _parse_myt_datetime(request_time)   or datetime.utcnow()
        completed_dt = _parse_myt_datetime(completed_time)
        if wo_stat == "completed" and not completed_dt:
            completed_dt = created_dt

        try:
            actual_hours_val = float(actual_hours) if actual_hours not in ("", "—") else None
        except ValueError:
            actual_hours_val = None

        if not wo_number:
            next_seq += 1
            wo_number = f"WO-{created_dt.strftime('%Y%m')}-{next_seq:04d}"

        desc = "[IMPORTED HISTORICAL RECORD]\n"
        if operator or category or problem:
            desc += (
                f"[OPERATOR REQUEST]\n"
                f"Submitted by  : {operator or '—'}\n"
                f"Category      : {category or '—'}\n"
                f"Problem       : {problem or '—'}\n"
            )
        if root_cause or action_taken or completed_by or wo_stat == "completed":
            stamp = (completed_dt + MY_TZ_OFFSET).strftime("%d %b %Y %H:%M") if completed_dt else "—"
            desc += (
                f"\n[COMPLETION — {stamp}]\n"
                f"Completed by  : {completed_by or '—'}\n"
                f"Root cause    : {root_cause or '—'}\n"
                f"Actions taken : {action_taken or '—'}\n"
            )
            if actual_hours_val is not None:
                desc += f"Actual hours  : {actual_hours_val}h\n"
            desc += f"Downtime type : {'Non affected' if downtime_type.startswith('non') else 'Affected'}\n"

        wo = WorkOrder(
            wo_number         = wo_number,
            asset_id          = asset.id,
            type              = WorkOrderType(wo_type),
            priority          = Priority(wo_pri),
            status            = WorkOrderStatus(wo_stat),
            title             = f"[Imported] {category or asset.name}",
            description       = desc,
            actual_hours      = actual_hours_val,
            affected_downtime = not downtime_type.startswith("non"),
            created_by        = current_user.id,
            created_at        = created_dt,
            completed_at      = completed_dt,
        )
        db.add(wo)
        existing_wo_numbers.add(wo_number)
        inserted += 1

    await db.flush()

    return {
        "success":         True,
        "inserted":        inserted,
        "skipped":         len(skipped),
        "skipped_details": skipped,
        "message":         f"Import complete: {inserted} work order(s) inserted, {len(skipped)} skipped.",
    }


# ─────────────────────────────────────────────────────────────────────────────
# HISTORY EXPORT  GET /data/export/history.xlsx
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/export/history.xlsx")
async def export_history(
    status:    Optional[str] = None,
    date_from: Optional[str] = None,
    date_to:   Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """Export work order history to a formatted Excel file."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(500, "openpyxl not installed")

    # Query
    q = (
        select(WorkOrder)
        .options(
            selectinload(WorkOrder.asset),
            selectinload(WorkOrder.assignments).selectinload(TaskAssignment.user),
        )
        .where(WorkOrder.is_deleted == False)
        .order_by(WorkOrder.created_at.desc())
    )
    if status:
        q = q.where(WorkOrder.status == status)
    if date_from:
        try:
            q = q.where(WorkOrder.created_at >= datetime.strptime(date_from, "%Y-%m-%d"))
        except ValueError:
            pass
    if date_to:
        try:
            dt = datetime.strptime(date_to, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
            q = q.where(WorkOrder.created_at <= dt)
        except ValueError:
            pass

    wos = (await db.execute(q)).scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Work Order History"

    thin   = Side(style="thin",   color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Title
    ws.merge_cells("A1:Q1")
    ws["A1"] = "CMMS Pro — Work Order History Report"
    ws["A1"].font      = Font(name="Arial", bold=True, size=13, color="1E3A5F")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].fill      = PatternFill("solid", start_color="EBF3FB")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:Q2")
    ws["A2"] = f"Generated: {datetime.utcnow().strftime('%d %b %Y %H:%M')} UTC   |   Total: {len(wos)} records"
    ws["A2"].font      = Font(name="Arial", italic=True, size=10, color="5A6A7A")
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 16

    ws.row_dimensions[3].height = 6  # spacer

    headers = [
        "WO Number","Machine","Asset Code","Location","Problem Category","Problem Description",
        "Type","Priority","Status","Operator",
        "Root Cause Analysis","Action Taken","Completed By",
        "Request Time (MYT)","Completed Time (MYT)",
        "Downtime / Actual Hours","Downtime Type",
    ]
    hdr_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    hdr_fill = PatternFill("solid", start_color="1E3A5F")
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.font = hdr_font; cell.fill = hdr_fill; cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[4].height = 26

    status_colors = {
        "completed":   "DCFCE7",
        "open":        "DBEAFE",
        "in_progress": "EDE9FE",
        "on_hold":     "FFEDD5",
        "cancelled":   "F1F5F9",
    }

    def parse_field(desc, key):
        for line in (desc or "").splitlines():
            if line.strip().startswith(key):
                parts = line.split(":", 1)
                return parts[1].strip() if len(parts) == 2 else ""
        return ""

    def downtime(wo):
        if wo.actual_hours:
            return round(wo.actual_hours, 2)
        if wo.completed_at and wo.created_at:
            return round((wo.completed_at - wo.created_at).total_seconds() / 3600, 2)
        if wo.created_at:
            return round((datetime.utcnow() - wo.created_at).total_seconds() / 3600, 2)
        return ""

    def assigned_technicians(wo):
        """All technicians who worked this job — not just whichever one
        happened to click "Complete". Prefers the TaskAssignment rows (the
        normal assign flow always writes these); falls back to parsing the
        "Assigned to" line in the description for older/imported records
        that predate — or bypassed — that flow, mirroring what the app's
        own UI already does when displaying assignees on a work order card."""
        names = []
        for a in sorted(wo.assignments or [], key=lambda a: a.assigned_at or datetime.min):
            name = a.user.name if a.user else None
            if name and name not in names:
                names.append(name)
        if names:
            return ", ".join(names)
        m = re.search(r"Assigned to\s*:\s*([^\n]+)", wo.description or "")
        if not m:
            return ""
        for part in m.group(1).split(","):
            name = re.sub(r"\s*\([^)]*\)", "", part).strip()
            if name and name not in names:
                names.append(name)
        return ", ".join(names)

    data_font = Font(name="Arial", size=10)

    for r, wo in enumerate(wos, 5):
        fill = PatternFill("solid", start_color=status_colors.get(wo.status.value if wo.status else "", "FFFFFF"))
        row_data = [
            wo.wo_number,
            wo.asset.name        if wo.asset else "—",
            wo.asset.asset_code  if wo.asset else "—",
            wo.asset.location    if wo.asset else "—",
            parse_field(wo.description, "Category") or "—",
            parse_field(wo.description, "Problem")  or "—",
            wo.type.value        if wo.type     else "—",
            wo.priority.value    if wo.priority else "—",
            wo.status.value      if wo.status   else "—",
            parse_field(wo.description, "Submitted by")  or "—",
            parse_field(wo.description, "Root cause")    or "—",
            parse_field(wo.description, "Actions taken") or "—",
            assigned_technicians(wo) or parse_field(wo.description, "Completed by") or "—",
            (wo.created_at + MY_TZ_OFFSET).strftime("%d %b %Y %H:%M")   if wo.created_at   else "—",
            (wo.completed_at + MY_TZ_OFFSET).strftime("%d %b %Y %H:%M") if wo.completed_at else "—",
            downtime(wo),
            "Affected" if getattr(wo, "affected_downtime", True) else "Non affected",
        ]
        for c, v in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = data_font; cell.fill = fill; cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=(c in (6, 11, 12)))
        ws.row_dimensions[r].height = 18

    # Column widths
    for i, w in enumerate([16,22,14,20,18,34,12,12,14,18,22,30,18,18,18,18,14], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A5"

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2.column_dimensions["A"].width = 26
    ws2.column_dimensions["B"].width = 14

    ws2.merge_cells("A1:B1")
    ws2["A1"] = "Summary"
    ws2["A1"].font      = Font(name="Arial", bold=True, size=13, color="1E3A5F")
    ws2["A1"].alignment = Alignment(horizontal="center")
    ws2["A1"].fill      = PatternFill("solid", start_color="EBF3FB")
    ws2.row_dimensions[1].height = 28

    from collections import Counter
    sc = Counter(wo.status.value   for wo in wos)
    pc = Counter(wo.priority.value for wo in wos)

    rows = [
        ("",""),("Total Work Orders", len(wos)),("",""),
        ("BY STATUS",""),
    ] + [(k.replace("_"," ").title(), v) for k,v in sc.items()] + [
        ("",""),("BY PRIORITY",""),
    ] + [(k.title(), v) for k,v in pc.items()]

    for r, (k, v) in enumerate(rows, 2):
        if k in ("BY STATUS","BY PRIORITY"):
            ws2.cell(row=r,column=1,value=k).font = Font(name="Arial",bold=True,size=11,color="1E3A5F")
            ws2.cell(row=r,column=1).fill = PatternFill("solid",start_color="EBF3FB")
            ws2.merge_cells(f"A{r}:B{r}")
        else:
            ws2.cell(row=r,column=1,value=k).font = Font(name="Arial",size=10,bold=(k=="Total Work Orders"))
            ws2.cell(row=r,column=2,value=v).font = Font(name="Arial",size=10)
        ws2.row_dimensions[r].height = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"WO_History_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )