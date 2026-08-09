"""
Work Orders Router — photos stored as compressed JPEGs on the persistent
disk (see photo_storage.py), referenced from work_order_photos rows.
"""
import re
import io
import os
import base64
from uuid import UUID
from datetime import datetime, timedelta
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy import select, func, delete
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload
from pydantic import BaseModel
from typing import Optional

from db import get_db
from models import WorkOrder, WorkOrderPhoto, PartsUsed, SparePart, User, WorkOrderStatus, UserRole, Asset, TaskAssignment, TaskStatus
from auth import get_current_user, forbid_viewer
from websocket_manager import ws_manager
from routers.analysis import working_hours_between, _wo_downtime_hours
from routers.stock import _log_movement, _ensure_table as _ensure_stock_table
from stock_status import compute_stock_status
from photo_storage import save_photo, PHOTOS_DIR
from wo_numbering import insert_with_unique_wo_number

router = APIRouter(prefix="/work-orders", tags=["work_orders"])
ROOM = "work_orders"


# ── Schemas ────────────────────────────────────────────────────────────────

class PhotoEntry(BaseModel):
    filename: str
    data: str   # base64 data URL
    size: int = 0

class CompletionForm(BaseModel):
    # root_cause and affected_downtime are omitted by the frontend when
    # completing a maintenance TASK (planned work, affected_downtime already
    # False) — those fields only apply to breakdown work orders.
    root_cause:        Optional[str] = None
    actions_taken:     str
    actual_hours:      float
    affected_downtime: Optional[bool] = None
    remarks:           Optional[str] = None
    parts_used:        list[dict] = []   # [{spare_part_id, quantity_used}]
    completion_photos: list[PhotoEntry] = []

class AssignRequest(BaseModel):
    user_id:  Optional[str] = None
    user_ids: list[str] = []
    priority: Optional[str] = None
    due_date: Optional[str] = None
    notes:    Optional[str] = None

class HoldRequest(BaseModel):
    reason: str

class WOCreate(BaseModel):
    asset_id:         str
    type:             str = "corrective"
    priority:         str = "medium"
    title:            str
    description:      Optional[str] = None
    due_date:         Optional[str] = None
    estimated_hours:  Optional[float] = None
    affected_downtime: bool = True

class WOUpdate(BaseModel):
    status:           Optional[str] = None
    priority:         Optional[str] = None
    title:            Optional[str] = None
    description:      Optional[str] = None
    due_date:         Optional[str] = None
    actual_hours:     Optional[float] = None
    affected_downtime: Optional[bool] = None
    assigned_to_user: Optional[str] = None
    asset_id:         Optional[str] = None


# ── Photo helpers ──────────────────────────────────────────────────────────

def _embed_photos(photos: list[PhotoEntry], tag: str, max_kb: int = 250) -> str:
    """Embed photos as base64 blocks with a section tag."""
    parts = []
    for p in photos[:3]:
        try:
            data_url = p.data if p.data.startswith("data:") else f"data:image/jpeg;base64,{p.data}"
            raw = data_url.split(",", 1)[1] if "," in data_url else data_url
            if len(raw) * 0.75 / 1024 > max_kb:
                parts.append(f"[PHOTO:{p.filename}|TOO_LARGE]")
                continue
            parts.append(f"[PHOTO:{p.filename}|{data_url}]")
        except Exception:
            pass
    if not parts:
        return ""
    return f"\n[{tag}]\n" + "\n".join(parts) + f"\n[/{tag}]\n"


# ── WO dict serializer ─────────────────────────────────────────────────────

def _wo_dict(wo: WorkOrder) -> dict:
    return {
        "id":           str(wo.id),
        "wo_number":    wo.wo_number,
        "title":        wo.title,
        "description":  wo.description,
        "type":         wo.type.value    if wo.type     else None,
        "priority":     wo.priority.value if wo.priority else None,
        "status":       wo.status.value  if wo.status   else None,
        "actual_hours": wo.actual_hours,
        "estimated_hours": wo.estimated_hours,
        "affected_downtime": bool(getattr(wo, "affected_downtime", True)),
        "hold_started_at": wo.hold_started_at.isoformat() if getattr(wo, "hold_started_at", None) else None,
        "held_hours":   getattr(wo, "held_hours", 0.0) or 0.0,
        "due_date":     wo.due_date.isoformat()     if wo.due_date     else None,
        "created_at":   wo.created_at.isoformat()   if wo.created_at   else None,
        "updated_at":   wo.updated_at.isoformat()   if wo.updated_at   else None,
        "completed_at": wo.completed_at.isoformat() if wo.completed_at else None,
        "is_deleted":   bool(getattr(wo, "is_deleted", False)),
        "deleted_at":   wo.deleted_at.isoformat()   if getattr(wo, "deleted_at", None)   else None,
        "deleted_by":   getattr(wo, "deleted_by", None),
        "restored_at":  wo.restored_at.isoformat()  if getattr(wo, "restored_at", None)  else None,
        "restored_by":  getattr(wo, "restored_by", None),
        # Only short URL strings here, never the image bytes themselves —
        # this is what keeps list/analysis responses small regardless of
        # how many photos a work order has. The browser fetches the actual
        # thumbnail/full image bytes separately (and only when needed) via
        # these /photos/... URLs, streamed straight from disk.
        "photos": [
            {
                "id":        str(ph.id),
                "kind":      ph.kind,
                "filename":  ph.filename,
                "thumb_url": f"/photos/{ph.thumb_path}",
                "full_url":  f"/photos/{ph.full_path}",
            }
            for ph in (wo.photos or [])
        ],
        "assigned_users": [
            {
                "id": str(a.user.id),
                "name": a.user.name,
                "email": a.user.email,
                "role": a.user.role.value if a.user.role else None,
                "assignment_status": a.status.value if a.status else None,
            }
            for a in (wo.assignments or [])
            if a.user and getattr(a.user, "is_active", True)
        ],
        "asset": {
            "id":         str(wo.asset.id),
            "name":       wo.asset.name,
            "asset_code": wo.asset.asset_code,
            "location":   wo.asset.location,
        } if wo.asset else None,
    }


async def _get_wo(wo_id: UUID, db: AsyncSession) -> WorkOrder:
    result = await db.execute(
        select(WorkOrder).where(WorkOrder.id == wo_id)
        .options(
            selectinload(WorkOrder.asset),
            selectinload(WorkOrder.assignments).selectinload(TaskAssignment.user),
            selectinload(WorkOrder.photos),
        )
    )
    wo = result.scalar_one_or_none()
    if not wo:
        raise HTTPException(404, "Work order not found")
    return wo


def _wo_number(seq: int) -> str:
    return f"WO-{datetime.utcnow().strftime('%Y%m')}-{seq:04d}"


# All datetime columns (created_at, completed_at, deleted_at, etc.) stay in
# UTC in the database — the frontend converts those to the browser's local
# time when displaying them. But the timestamps stamped directly into the
# free-text description (e.g. "[COMPLETION — 15 Jul 2026 14:30]") are plain
# strings, not real datetimes, so nothing converts them client-side. This
# formats "now" in Malaysia time (UTC+8, no DST) for those description
# stamps specifically, so they read correctly wherever they're displayed.
def _my_now_str(fmt: str = "%d %b %Y %H:%M") -> str:
    return (datetime.utcnow() + timedelta(hours=8)).strftime(fmt)


def _apply_hold_transition(wo: WorkOrder, new_status: WorkOrderStatus) -> None:
    """
    Keeps hold_started_at / held_hours in sync whenever a work order's
    status is about to change. Call this BEFORE assigning wo.status.

    - Entering on_hold (from anything else): stamp hold_started_at = now,
      so downtime stops accumulating from this point.
    - Leaving on_hold (to anything else): fold the working hours spent in
      that hold window into held_hours and clear hold_started_at, so
      downtime resumes counting from where it left off.
    """
    now = datetime.utcnow()
    was_on_hold = wo.status == WorkOrderStatus.on_hold
    going_on_hold = new_status == WorkOrderStatus.on_hold

    if going_on_hold and not was_on_hold:
        wo.hold_started_at = now
    elif was_on_hold and not going_on_hold and wo.hold_started_at:
        wo.held_hours = (wo.held_hours or 0.0) + working_hours_between(wo.hold_started_at, now)
        wo.hold_started_at = None


# ── CRUD ───────────────────────────────────────────────────────────────────

@router.get("/")
async def list_work_orders(
    skip: int = 0, limit: int = 100,
    status: str | None = None,
    priority: str | None = None,
    asset_id: UUID | None = None,
    my_jobs_only: bool = False,
    deleted_only: bool = False,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    q = select(WorkOrder).options(
        selectinload(WorkOrder.asset),
        selectinload(WorkOrder.assignments).selectinload(TaskAssignment.user),
        selectinload(WorkOrder.photos),
    )
    # Soft-deleted work orders are hidden everywhere by default. Only admins
    # can see the "Deleted Work Orders" trash view (deleted_only=True) —
    # everyone else always gets the normal, non-deleted list regardless of
    # what they pass, so a deleted WO never resurfaces on the Alert Board,
    # technician's board, or anyone else's History Log by mistake.
    if deleted_only and current_user.role == UserRole.admin:
        q = q.where(WorkOrder.is_deleted == True)
    else:
        q = q.where(WorkOrder.is_deleted == False)
    if status:   q = q.where(WorkOrder.status == status)
    if priority: q = q.where(WorkOrder.priority == priority)
    if asset_id: q = q.where(WorkOrder.asset_id == asset_id)

    # Explicit opt-in scoping — used by the technician Work Order Board so they
    # only see jobs assigned to them. Admin/manager are never restricted, even
    # if this flag is passed, since they're allowed full visibility everywhere.
    if my_jobs_only and current_user.role.value not in ("admin", "manager"):
        q = q.where(
            WorkOrder.assignments.any(TaskAssignment.user_id == current_user.id)
        )

    q = q.offset(skip).limit(limit).order_by(WorkOrder.created_at.desc())
    result = await db.execute(q)
    return [_wo_dict(w) for w in result.scalars().all()]


@router.get("/{wo_id}")
async def get_work_order(wo_id: UUID, db: AsyncSession = Depends(get_db), _: User = Depends(get_current_user)):
    return _wo_dict(await _get_wo(wo_id, db))


@router.post("/", status_code=201)
async def create_work_order(
    body: WOCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(forbid_viewer),
):
    from models import WorkOrderType, Priority
    created_by_id = current_user.id  # snapshot now — a wo_number collision
    # retry (see wo_numbering.py) rolls back and would expire current_user,
    # so reading current_user.id again inside a re-invoked build() closure
    # would trigger an async-unsafe lazy reload.

    async def build(wo_number: str) -> WorkOrder:
        wo = WorkOrder(
            wo_number       = wo_number,
            asset_id        = UUID(body.asset_id),
            type            = WorkOrderType(body.type),
            priority        = Priority(body.priority),
            status          = WorkOrderStatus.open,
            title           = body.title,
            description     = body.description,
            estimated_hours = body.estimated_hours,
            affected_downtime = body.affected_downtime,
            created_by      = created_by_id,
        )
        if body.due_date:
            wo.due_date = datetime.fromisoformat(body.due_date.replace("Z",""))
        db.add(wo)
        return wo

    wo = await insert_with_unique_wo_number(db, build, prefix="MT" if body.type=="preventive" else "WO")
    await db.refresh(wo)
    await ws_manager.broadcast_event(ROOM, "work_order.created", {
        "id": str(wo.id), "wo_number": wo.wo_number, "title": wo.title,
        "priority": wo.priority.value, "status": wo.status.value,
    })
    return _wo_dict(await _get_wo(wo.id, db))


@router.patch("/{wo_id}")
async def update_work_order(
    wo_id: UUID, body: WOUpdate,
    db: AsyncSession = Depends(get_db), current_user: User = Depends(forbid_viewer),
):
    from models import Priority
    wo = await _get_wo(wo_id, db)
    if wo.is_deleted:
        raise HTTPException(400, "This work order has been deleted — restore it first")
    if body.status:
        new_status = WorkOrderStatus(body.status)
        _apply_hold_transition(wo, new_status)
        wo.status = new_status
        if body.status == "completed" and not wo.completed_at:
            wo.completed_at = datetime.utcnow()
    if body.priority:
        wo.priority = Priority(body.priority)
    if body.title:
        wo.title = body.title
    if body.description is not None:
        wo.description = body.description
    if body.due_date:
        wo.due_date = datetime.fromisoformat(body.due_date.replace("Z",""))
    if body.actual_hours is not None:
        wo.actual_hours = body.actual_hours
    if body.affected_downtime is not None:
        wo.affected_downtime = body.affected_downtime
    if body.assigned_to_user:
        wo.assigned_to_user = UUID(body.assigned_to_user)
    if body.asset_id:
        # Operators sometimes report the wrong machine on a request — let an
        # admin correct which asset a work order is attached to. Restricted
        # to admins since this changes downtime/analysis attribution.
        if current_user.role != UserRole.admin:
            raise HTTPException(403, "Only admins can change a work order's machine")
        new_asset = await db.get(Asset, UUID(body.asset_id))
        if not new_asset:
            raise HTTPException(404, "Asset not found")
        wo.asset_id = new_asset.id
    wo.updated_at = datetime.utcnow()
    await db.flush()
    await ws_manager.broadcast_event(ROOM, "work_order.updated", {
        "id": str(wo.id), "wo_number": wo.wo_number,
        "status": wo.status.value, "priority": wo.priority.value,
    })
    return _wo_dict(await _get_wo(wo_id, db))


@router.delete("/{wo_id}")
async def delete_work_order(
    wo_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(forbid_viewer),
):
    """
    Soft-deletes a work order — admin only. The record isn't removed from
    the database, just hidden from every normal view (Alert Board,
    technician's board, History Log, reports) and marked with who deleted
    it and when, so it can be reviewed and restored later from History
    Log's "Deleted Work Orders" view.
    """
    if current_user.role != UserRole.admin:
        raise HTTPException(403, "Only admins can delete work orders")
    wo = await _get_wo(wo_id, db)
    if wo.is_deleted:
        raise HTTPException(400, "Work order is already deleted")
    wo.is_deleted = True
    wo.deleted_at = datetime.utcnow()
    wo.deleted_by = current_user.name
    await db.flush()
    await ws_manager.broadcast_event(ROOM, "work_order.deleted", {
        "id": str(wo_id), "wo_number": wo.wo_number, "deleted_by": current_user.name,
    })
    return _wo_dict(await _get_wo(wo_id, db))


@router.post("/{wo_id}/restore")
async def restore_work_order(
    wo_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(forbid_viewer),
):
    """Reverses a soft-delete — admin only. Brings the work order back into
    every normal view exactly as it was before deletion."""
    if current_user.role != UserRole.admin:
        raise HTTPException(403, "Only admins can restore work orders")
    wo = await _get_wo(wo_id, db)
    if not wo.is_deleted:
        raise HTTPException(400, "Work order is not deleted")
    wo.is_deleted = False
    wo.restored_at = datetime.utcnow()
    wo.restored_by = current_user.name
    await db.flush()
    await ws_manager.broadcast_event(ROOM, "work_order.restored", {
        "id": str(wo_id), "wo_number": wo.wo_number, "restored_by": current_user.name,
    })
    return _wo_dict(await _get_wo(wo_id, db))


# ── Assign to technician ───────────────────────────────────────────────────

@router.post("/{wo_id}/assign")
async def assign_to_technician(
    wo_id: UUID, body: AssignRequest,
    db: AsyncSession = Depends(get_db), current_user: User = Depends(forbid_viewer),
):
    from models import Priority

    # Collect all requested user IDs (supports legacy single user_id + new user_ids list)
    raw_ids: list[str] = list(body.user_ids or [])
    if body.user_id and body.user_id not in raw_ids:
        raw_ids.append(body.user_id)
    if not raw_ids:
        raise HTTPException(400, "At least one technician must be selected")

    wo = await _get_wo(wo_id, db)
    if wo.is_deleted:
        raise HTTPException(400, "This work order has been deleted — restore it first")

    techs: list[User] = []
    for uid in raw_ids:
        tech = (await db.execute(select(User).where(User.id == UUID(uid)))).scalar_one_or_none()
        if not tech:
            raise HTTPException(404, f"User {uid} not found")
        techs.append(tech)

    # Replace existing assignments with the new set
    await db.execute(delete(TaskAssignment).where(TaskAssignment.work_order_id == wo_id))
    for tech in techs:
        db.add(TaskAssignment(
            work_order_id=wo_id,
            user_id=tech.id,
            status=TaskStatus.pending,
            notes=body.notes,
        ))

    _apply_hold_transition(wo, WorkOrderStatus.in_progress)
    wo.status = WorkOrderStatus.in_progress
    wo.updated_at = datetime.utcnow()
    if body.priority:
        wo.priority = Priority(body.priority)
    if body.due_date:
        wo.due_date = datetime.fromisoformat(body.due_date.replace("Z",""))

    tech_names = ", ".join(f"{t.name} ({t.email})" for t in techs)
    note = (
        f"\n[ASSIGNED — {_my_now_str()}]\n"
        f"Assigned to   : {tech_names}\n"
    )
    if body.notes:
        note += f"Notes         : {body.notes}\n"
    wo.description = (wo.description or "") + note

    await db.flush()
    await ws_manager.broadcast_event(ROOM, "work_order.assigned", {
        "id": str(wo.id), "wo_number": wo.wo_number,
        "assigned_to": tech_names,
        "assigned_user_ids": [str(t.id) for t in techs],
    })
    return _wo_dict(await _get_wo(wo_id, db))


# ── Hold / resume ───────────────────────────────────────────────────────────

@router.post("/{wo_id}/hold")
async def hold_work_order(
    wo_id: UUID, body: HoldRequest,
    db: AsyncSession = Depends(get_db), current_user: User = Depends(forbid_viewer),
):
    """
    Puts a work order on hold with a required remark explaining why —
    e.g. waiting on a spare part, waiting on production to stop the line,
    escalated to a vendor. The remark is recorded in the description's
    audit trail, and the downtime clock freezes until the WO is resumed
    or completed (see _apply_hold_transition).
    """
    if not body.reason or not body.reason.strip():
        raise HTTPException(400, "Please provide a reason for the hold")

    wo = await _get_wo(wo_id, db)
    if wo.is_deleted:
        raise HTTPException(400, "This work order has been deleted — restore it first")
    if wo.status == WorkOrderStatus.completed:
        raise HTTPException(400, "Cannot put a completed work order on hold")
    if wo.status == WorkOrderStatus.on_hold:
        raise HTTPException(400, "Work order is already on hold")

    reason = body.reason.strip()
    _apply_hold_transition(wo, WorkOrderStatus.on_hold)
    wo.status     = WorkOrderStatus.on_hold
    wo.updated_at = datetime.utcnow()
    wo.description = (wo.description or "") + (
        f"\n[ON HOLD — {_my_now_str()}]\n"
        f"Held by      : {current_user.name}\n"
        f"Reason       : {reason}\n"
    )

    await db.flush()
    await ws_manager.broadcast_event(ROOM, "work_order.on_hold", {
        "id": str(wo.id), "wo_number": wo.wo_number,
        "held_by": current_user.name, "reason": reason,
    })
    return _wo_dict(await _get_wo(wo_id, db))


@router.post("/{wo_id}/resume")
async def resume_work_order(
    wo_id: UUID,
    db: AsyncSession = Depends(get_db), current_user: User = Depends(forbid_viewer),
):
    """Takes a work order off hold, back to in_progress, resuming the downtime clock."""
    wo = await _get_wo(wo_id, db)
    if wo.is_deleted:
        raise HTTPException(400, "This work order has been deleted — restore it first")
    if wo.status != WorkOrderStatus.on_hold:
        raise HTTPException(400, "Work order is not on hold")

    _apply_hold_transition(wo, WorkOrderStatus.in_progress)
    wo.status     = WorkOrderStatus.in_progress
    wo.updated_at = datetime.utcnow()
    wo.description = (wo.description or "") + (
        f"\n[RESUMED — {_my_now_str()}]\n"
        f"Resumed by   : {current_user.name}\n"
    )

    await db.flush()
    await ws_manager.broadcast_event(ROOM, "work_order.resumed", {
        "id": str(wo.id), "wo_number": wo.wo_number, "resumed_by": current_user.name,
    })
    return _wo_dict(await _get_wo(wo_id, db))


# ── Completion form ────────────────────────────────────────────────────────

@router.post("/{wo_id}/complete")
async def complete_work_order(
    wo_id: UUID, body: CompletionForm,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(forbid_viewer),
):
    wo = await _get_wo(wo_id, db)
    if wo.is_deleted:
        raise HTTPException(400, "This work order has been deleted — restore it first")
    if wo.status == WorkOrderStatus.completed:
        raise HTTPException(400, "Work order already completed")

    # A maintenance task is planned work (no production breakdown behind
    # it) — it never asks for a root cause or downtime type, so don't
    # require them here either. Real work orders still must supply both.
    is_task = wo.affected_downtime is False
    if not is_task and not body.root_cause:
        raise HTTPException(400, "Root cause is required to complete a work order")

    _apply_hold_transition(wo, WorkOrderStatus.completed)
    wo.status       = WorkOrderStatus.completed
    wo.completed_at = datetime.utcnow()
    wo.actual_hours = body.actual_hours
    if body.affected_downtime is not None:
        wo.affected_downtime = body.affected_downtime
    wo.updated_at   = datetime.utcnow()

    comp = (
        f"\n[COMPLETION — {_my_now_str()}]\n"
        f"Completed by  : {current_user.name}\n"
    )
    if body.root_cause:
        comp += f"Root cause    : {body.root_cause}\n"
    comp += (
        f"Actions taken : {body.actions_taken}\n"
        f"Actual hours  : {body.actual_hours}h\n"
    )
    if body.affected_downtime is not None:
        comp += f"Downtime type : {'Affected' if body.affected_downtime else 'Non affected'}\n"
    if body.remarks:
        comp += f"Remarks       : {body.remarks}\n"

    wo.description = (wo.description or "") + comp

    # Save completion photos to disk (compressed full + thumbnail) instead
    # of embedding base64 in description — see photo_storage.py.
    for p in body.completion_photos[:3]:
        saved = save_photo(p.data, f"completion/{wo_id}")
        if saved:
            db.add(WorkOrderPhoto(
                work_order_id=wo_id, kind="completion", filename=p.filename,
                thumb_path=saved["thumb_path"], full_path=saved["full_path"],
            ))

    # Deduct parts from inventory
    if body.parts_used:
        await _ensure_stock_table(db)
    for entry in body.parts_used:
        part = (await db.execute(
            select(SparePart).where(SparePart.id == UUID(entry["spare_part_id"]))
        )).scalar_one_or_none()
        if not part:
            raise HTTPException(404, f"Part not found: {entry['spare_part_id']}")
        qty = int(entry.get("quantity_used", 1))
        if part.quantity_on_hand < qty:
            raise HTTPException(400, f"Insufficient stock for {part.name}. Available: {part.quantity_on_hand}")
        qty_before = part.quantity_on_hand
        part.quantity_on_hand -= qty
        db.add(PartsUsed(
            work_order_id = wo_id,
            spare_part_id = UUID(entry["spare_part_id"]),
            quantity_used = qty,
        ))
        await _log_movement(
            db, part, "stock_out", qty,
            qty_before, part.quantity_on_hand, current_user,
            reason="Used on work order completion", reference=wo.wo_number,
        )
        severity = compute_stock_status(part.quantity_on_hand, part.reorder_level, bool(getattr(part,"is_critical",False)))
        if severity != "ok":
            await ws_manager.broadcast_event("inventory", "inventory.low_stock", {
                "part_id": str(part.id), "part_name": part.name,
                "quantity_on_hand": part.quantity_on_hand,
                "stock_status": severity,
            })

    await db.flush()
    await ws_manager.broadcast_event(ROOM, "work_order.completed", {
        "id": str(wo_id), "wo_number": wo.wo_number,
        "completed_by": current_user.name, "actual_hours": body.actual_hours,
    })
    return _wo_dict(await _get_wo(wo_id, db))


# ── Parts availability ─────────────────────────────────────────────────────

@router.get("/{wo_id}/parts-availability")
async def parts_availability(
    wo_id: UUID,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_current_user),
):
    result = await db.execute(select(SparePart).order_by(SparePart.name))
    return [{
        "id":               str(p.id),
        "part_code":        p.part_code,
        "name":             p.name,
        "category":         p.category,
        "unit":             p.unit,
        "quantity_on_hand": p.quantity_on_hand,
        "reorder_level":    p.reorder_level,
        "is_available":     p.quantity_on_hand > 0,
        "is_low_stock":     compute_stock_status(p.quantity_on_hand, p.reorder_level, bool(getattr(p,"is_critical",False))) != "ok",
        "stock_status":     compute_stock_status(p.quantity_on_hand, p.reorder_level, bool(getattr(p,"is_critical",False))),
        "unit_cost":        p.unit_cost,
        "location":         p.location,
        "barcode":          getattr(p, "barcode", None),
    } for p in result.scalars().all()]


# ── MRR (Machine Repair Request) export ─────────────────────────────────────
# Fills the company's official MRR form (Doc No F/SMT-06) for a single work
# order, using the exact controlled-document template rather than
# recreating it from scratch, so it matches what's already printed/filed
# on paper. Everything gets pulled from data already on the work order —
# the free-text description tags written by requests.py (on submission)
# and the completion flow (on completion) — see those for the "Label :
# value" convention this parses.

_MRR_TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), "..", "templates", "MRR_form.xlsx")


def _extract_field(description: str, *labels: str) -> str:
    """Pulls a 'Label   : value' line out of a work order's free-text
    description. Tries each label in turn and returns the first match."""
    if not description:
        return ""
    for label in labels:
        m = re.search(rf"^{re.escape(label)}\s*:\s*(.+)$", description, re.MULTILINE)
        if m:
            return m.group(1).strip()
    return ""


@router.get("/{wo_id}/export/mrr")
async def export_mrr_form(
    wo_id: UUID,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_current_user),
):
    if not os.path.exists(_MRR_TEMPLATE_PATH):
        raise HTTPException(500, "MRR template is missing from the server.")
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Alignment, Font
        from openpyxl.drawing.image import Image as XLImage
        from PIL import Image as PILImage
    except ImportError:
        raise HTTPException(500, "openpyxl/Pillow not installed on server. Run: pip install openpyxl Pillow")

    wo = await _get_wo(wo_id, db)
    desc = wo.description or ""

    requested_by = _extract_field(desc, "Submitted by", "Requested by")
    category     = _extract_field(desc, "Category")
    problem      = _extract_field(desc, "Problem") or wo.title
    remarks_in   = _extract_field(desc, "Remarks")
    tested_by    = _extract_field(desc, "Completed by")
    root_cause   = _extract_field(desc, "Root cause")
    actions      = _extract_field(desc, "Actions taken")

    problem_lines = []
    if category: problem_lines.append(f"Category: {category}")
    problem_lines.append(problem)
    if remarks_in: problem_lines.append(f"Remarks: {remarks_in}")
    problem_text = "\n".join(problem_lines)

    corrective_lines = []
    if root_cause: corrective_lines.append(f"Root cause: {root_cause}")
    corrective_lines.append(actions or "")
    corrective_text = "\n".join(corrective_lines)

    divisor = wo.asset.downtime_divisor if wo.asset else 1.0
    downtime_hours = _wo_downtime_hours(wo.actual_hours, wo.created_at, wo.completed_at, wo.held_hours, divisor)

    # Timestamps are stored in UTC — shift to Malaysia time (UTC+8, no DST)
    # for display, matching the completion-note stamps elsewhere in the app.
    broke_down = (wo.created_at + timedelta(hours=8)) if wo.created_at else None
    fixed_up   = (wo.completed_at + timedelta(hours=8)) if wo.completed_at else None

    wb = load_workbook(_MRR_TEMPLATE_PATH)
    ws = wb["Repair Request"]

    ws["D6"] = wo.wo_number
    if broke_down:
        ws["B9"] = broke_down.strftime("%d %b %Y")
        ws["D9"] = broke_down.strftime("%H:%M")
    ws["B10"] = f"{wo.asset.name} ({wo.asset.asset_code})" if wo.asset else ""
    ws["D10"] = requested_by

    ws.merge_cells("B12:D17")
    ws["B12"] = problem_text
    ws["B12"].alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")

    ws.merge_cells("B20:D26")
    ws["B20"] = corrective_text
    ws["B20"].alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")

    if fixed_up:
        ws["B27"] = fixed_up.strftime("%d %b %Y")
        ws["D27"] = fixed_up.strftime("%H:%M")
    ws["B28"] = tested_by
    if downtime_hours:
        ws["D28"] = round(downtime_hours, 2)
    # B33/D33 ("Verified by requestor" / "Date") are left blank — that's a
    # physical signature line, signed after this is printed.

    # ── Photo Evidence sheet ──────────────────────────────────────────────
    # A second sheet, separate from the controlled "Repair Request" print
    # layout above (which stays byte-identical to the official form) —
    # operator photos from the initial report and technician photos from
    # completion, grouped and labeled so it's clear which is which.
    photos = list(wo.photos or [])
    if photos:
        ws2 = wb.create_sheet("Photos")
        ws2.column_dimensions["B"].width = 46
        ws2["B1"] = f"Photo Evidence — {wo.wo_number}"
        ws2["B1"].font = Font(bold=True, size=14)

        groups = [
            ("operator",   "📋 Reported by Operator"),
            ("completion", "🔧 Completed by Technician"),
        ]
        row = 3
        for kind_key, label in groups:
            group = [p for p in photos if p.kind == kind_key]
            if not group:
                continue
            ws2.cell(row=row, column=2, value=label).font = Font(bold=True, size=12)
            row += 2
            for ph in group:
                abs_path = os.path.join(PHOTOS_DIR, ph.full_path)
                if not os.path.exists(abs_path):
                    continue  # photo file missing on disk — skip rather than fail the export
                try:
                    with PILImage.open(abs_path) as im:
                        w, h = im.size
                    scale = min(1.0, 340 / max(w, h))
                    disp_w, disp_h = int(w * scale), int(h * scale)
                    xl_img = XLImage(abs_path)
                    xl_img.width, xl_img.height = disp_w, disp_h
                    ws2.add_image(xl_img, f"B{row}")
                    needed_rows = max(1, -(-disp_h // 20))  # ceil(px / ~20px per default row)
                    cap_row = row + needed_rows
                    caption = ph.filename or "photo"
                    if ph.created_at:
                        caption += "  ·  " + (ph.created_at + timedelta(hours=8)).strftime("%d %b %Y %H:%M")
                    ws2.cell(row=cap_row, column=2, value=caption).font = Font(italic=True, size=9, color="64748B")
                    row = cap_row + 2
                except Exception:
                    continue  # corrupt/unreadable image — skip rather than fail the export
            row += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"MRR_{wo.wo_number}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )