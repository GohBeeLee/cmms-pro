"""
Stock Router — barcode scan, stock in/out, history log
=======================================================
GET  /stock/scan?barcode=...      — lookup part by barcode or part_code
POST /stock/in                    — stock in (add qty)
POST /stock/out                   — stock out (remove qty)
GET  /stock/history               — full stock movement log
GET  /stock/history/{part_id}     — movements for one part
"""
import json
from datetime import datetime
from typing import Optional
from uuid import UUID, uuid4

from fastapi import APIRouter, Depends, HTTPException, Query
from sqlalchemy import select, desc, text
from sqlalchemy.ext.asyncio import AsyncSession
from pydantic import BaseModel

from db import get_db
from models import SparePart, User
from auth import get_current_user, forbid_viewer
from websocket_manager import ws_manager
from stock_status import compute_stock_status

router = APIRouter(prefix="/stock", tags=["stock"], dependencies=[Depends(forbid_viewer)])


def _require_admin_or_manager(user: User):
    if user.role.value not in ("admin", "manager"):
        raise HTTPException(403, "Only admins and managers can perform a stock take")


# ── Stock movement log stored in a simple JSON table ─────────────────────
# We store in a dedicated table via raw SQL so we don't need a migration.
# On first call, the table is created if it doesn't exist.

CREATE_TABLE = """
CREATE TABLE IF NOT EXISTS stock_movements (
    id          TEXT PRIMARY KEY,
    part_id     TEXT NOT NULL,
    part_code   TEXT NOT NULL,
    part_name   TEXT NOT NULL,
    movement    TEXT NOT NULL,
    quantity    INTEGER NOT NULL,
    qty_before  INTEGER NOT NULL,
    qty_after   INTEGER NOT NULL,
    reason      TEXT,
    reference   TEXT,
    performed_by_id   TEXT,
    performed_by_name TEXT,
    created_at  TEXT NOT NULL
)
"""


async def _ensure_table(db: AsyncSession):
    await db.execute(text(CREATE_TABLE))
    await db.commit()


async def _log_movement(
    db:             AsyncSession,
    part:           SparePart,
    movement:       str,          # "stock_in" | "stock_out" | "restock" | "adjustment"
    quantity:       int,
    qty_before:     int,
    qty_after:      int,
    user:           User,
    reason:         Optional[str] = None,
    reference:      Optional[str] = None,
):
    await db.execute(
        text("""
            INSERT INTO stock_movements
            (id, part_id, part_code, part_name, movement, quantity,
             qty_before, qty_after, reason, reference,
             performed_by_id, performed_by_name, created_at)
            VALUES (:id, :part_id, :part_code, :part_name, :movement, :quantity,
                    :qty_before, :qty_after, :reason, :reference,
                    :performed_by_id, :performed_by_name, :created_at)
        """),
        {
            "id":                 str(uuid4()),
            "part_id":            str(part.id),
            "part_code":          part.part_code,
            "part_name":          part.name,
            "movement":           movement,
            "quantity":           quantity,
            "qty_before":         qty_before,
            "qty_after":          qty_after,
            "reason":             reason,
            "reference":          reference,
            "performed_by_id":    str(user.id),
            "performed_by_name":  user.name,
            "created_at":         datetime.utcnow().isoformat(),
        }
    )


def _part_summary(p: SparePart) -> dict:
    qty   = p.quantity_on_hand or 0
    reord = p.reorder_level    or 0
    is_critical_flag = bool(getattr(p, "is_critical", False))
    severity = compute_stock_status(qty, reord, is_critical_flag)
    return {
        "id":               str(p.id),
        "part_code":        p.part_code,
        "name":             p.name,
        "category":         p.category,
        "description":      p.description,
        "quantity_on_hand": qty,
        "reorder_level":    reord,
        "unit_cost":        p.unit_cost,
        "location":         p.location,
        "barcode":          getattr(p, "barcode",       None),
        "used_on_asset":    getattr(p, "used_on_asset", None),
        "photo_thumb_url":  f"/photos/{p.photo_thumb_path}" if getattr(p, "photo_thumb_path", None) else None,
        "photo_full_url":   f"/photos/{p.photo_full_path}"  if getattr(p, "photo_full_path",  None) else None,
        "has_photo":        bool(getattr(p, "photo_thumb_path", None)),
        "last_stock_take_at": p.last_stock_take_at.isoformat() if getattr(p, "last_stock_take_at", None) else None,
        "last_stock_take_by": getattr(p, "last_stock_take_by", None),
        "unit":             p.unit or "pcs",
        "status":           "Topup" if qty <= reord else "Enough",
        "is_low_stock":     severity != "ok",
        "is_critical":      is_critical_flag,
        "stock_status":     severity,
    }


# ── Schemas ────────────────────────────────────────────────────────────────

class StockMoveIn(BaseModel):
    part_id:   str
    quantity:  int   = 1
    reason:    Optional[str] = None
    reference: Optional[str] = None   # e.g. PO number, delivery note

class StockMoveOut(BaseModel):
    part_id:   str
    quantity:  int   = 1
    reason:    Optional[str] = None
    reference: Optional[str] = None   # e.g. WO number

class StockTakeAdjust(BaseModel):
    part_id:          str
    new_quantity:     int
    reason:           Optional[str] = None   # e.g. "Physical count mismatch"
    # Optional detail corrections found during the physical count —
    # any field left as None is left unchanged.
    name:             Optional[str]   = None
    category:         Optional[str]   = None
    description:      Optional[str]   = None
    location:          Optional[str]   = None
    used_on_asset:    Optional[str]   = None
    unit_cost:        Optional[float] = None
    reorder_level:    Optional[int]   = None


# ══════════════════════════════════════════════════════════════════════════
# ROUTES
# ══════════════════════════════════════════════════════════════════════════

# ── Lookup part by barcode ─────────────────────────────────────────────────

@router.get("/scan")
async def scan_barcode(
    barcode: str = Query(..., description="Barcode or part_code to look up — accepts any characters including / and -"),
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """
    Look up a spare part by barcode OR part_code.
    Uses a query parameter (?barcode=...) instead of a path segment so that
    special characters like / - # are handled correctly without URL encoding
    ambiguity (a / in a path parameter breaks FastAPI routing entirely).
    Called when technician scans or types a barcode.
    """
    await _ensure_table(db)

    # Normalize: trim whitespace, try exact match first
    code = (barcode or "").strip()

    # Try barcode field first (new column)
    try:
        result = await db.execute(
            select(SparePart).where(SparePart.barcode == code)
        )
        part = result.scalar_one_or_none()
    except Exception:
        part = None

    # Fallback: match part_code
    if not part:
        result = await db.execute(
            select(SparePart).where(SparePart.part_code == code)
        )
        part = result.scalar_one_or_none()

    if not part:
        raise HTTPException(
            404,
            f"No part found for barcode '{barcode}'. "
            "Make sure the barcode is set in the Inventory page."
        )

    # Fetch last 5 movements for this part
    try:
        rows = (await db.execute(
            text("""
                SELECT movement, quantity, qty_before, qty_after,
                       performed_by_name, reason, created_at
                FROM stock_movements
                WHERE part_id = :pid
                ORDER BY created_at DESC LIMIT 5
            """),
            {"pid": str(part.id)}
        )).fetchall()
        recent = [dict(r._mapping) for r in rows]
    except Exception:
        recent = []

    return {**_part_summary(part), "recent_movements": recent}


# ── Stock In ───────────────────────────────────────────────────────────────

@router.post("/in")
async def stock_in(
    body: StockMoveIn,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Add stock — technician received / restocked items."""
    await _ensure_table(db)

    if body.quantity < 1:
        raise HTTPException(400, "Quantity must be at least 1")

    part = (await db.execute(
        select(SparePart).where(SparePart.id == UUID(body.part_id))
    )).scalar_one_or_none()
    if not part:
        raise HTTPException(404, "Part not found")

    qty_before = part.quantity_on_hand or 0
    qty_after  = qty_before + body.quantity
    part.quantity_on_hand = qty_after
    part.updated_at       = datetime.utcnow()

    await _log_movement(
        db, part, "stock_in", body.quantity,
        qty_before, qty_after, current_user,
        reason=body.reason, reference=body.reference,
    )
    await db.flush()

    await ws_manager.broadcast_event("inventory", "inventory.stock_in", {
        "part_id":   str(part.id),
        "part_name": part.name,
        "quantity":  body.quantity,
        "new_qty":   qty_after,
        "by":        current_user.name,
    })

    return {
        "success":    True,
        "message":    f"Stock In: +{body.quantity} {part.unit or 'pcs'} added to {part.name}",
        "part":       _part_summary(part),
        "qty_before": qty_before,
        "qty_after":  qty_after,
    }


# ── Stock Out ──────────────────────────────────────────────────────────────

@router.post("/out")
async def stock_out(
    body: StockMoveOut,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Remove stock — technician used items on a job."""
    await _ensure_table(db)

    if body.quantity < 1:
        raise HTTPException(400, "Quantity must be at least 1")

    part = (await db.execute(
        select(SparePart).where(SparePart.id == UUID(body.part_id))
    )).scalar_one_or_none()
    if not part:
        raise HTTPException(404, "Part not found")

    qty_before = part.quantity_on_hand or 0
    if qty_before < body.quantity:
        raise HTTPException(
            400,
            f"Insufficient stock for {part.name}. "
            f"Available: {qty_before} {part.unit or 'pcs'}, "
            f"Requested: {body.quantity}"
        )

    qty_after = qty_before - body.quantity
    part.quantity_on_hand = qty_after
    part.updated_at       = datetime.utcnow()

    await _log_movement(
        db, part, "stock_out", body.quantity,
        qty_before, qty_after, current_user,
        reason=body.reason, reference=body.reference,
    )
    await db.flush()

    # Low stock alert
    severity = compute_stock_status(qty_after, part.reorder_level or 0, bool(getattr(part,"is_critical",False)))
    if severity != "ok":
        await ws_manager.broadcast_event("inventory", "inventory.low_stock", {
            "part_id":      str(part.id),
            "part_name":    part.name,
            "quantity":     qty_after,
            "reorder_level":part.reorder_level,
            "stock_status": severity,
        })

    await ws_manager.broadcast_event("inventory", "inventory.stock_out", {
        "part_id":   str(part.id),
        "part_name": part.name,
        "quantity":  body.quantity,
        "new_qty":   qty_after,
        "by":        current_user.name,
    })

    return {
        "success":    True,
        "message":    f"Stock Out: −{body.quantity} {part.unit or 'pcs'} removed from {part.name}",
        "part":       _part_summary(part),
        "qty_before": qty_before,
        "qty_after":  qty_after,
        "low_stock":  severity != "ok",
        "stock_status": severity,
    }


# ── Stock Take (physical count adjustment) ──────────────────────────────────

@router.post("/adjust")
async def stock_take_adjust(
    body: StockTakeAdjust,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Admin/manager only. Used during a physical stock take: scan a part,
    compare its system quantity (and details) against what's actually on
    the shelf, then correct anything that doesn't match. Always logged as
    an "adjustment" movement so there's a full audit trail of every
    correction — who made it, when, what changed, and why.
    """
    _require_admin_or_manager(current_user)
    await _ensure_table(db)

    part = (await db.execute(
        select(SparePart).where(SparePart.id == UUID(body.part_id))
    )).scalar_one_or_none()
    if not part:
        raise HTTPException(404, "Part not found")

    qty_before = part.quantity_on_hand or 0
    qty_after  = body.new_quantity
    if qty_after < 0:
        raise HTTPException(400, "Quantity cannot be negative")

    field_changes = []

    if qty_after != qty_before:
        part.quantity_on_hand = qty_after
        field_changes.append(f"Qty {qty_before} -> {qty_after}")

    # Optional detail corrections — only touch fields actually supplied
    detail_fields = {
        "name": body.name, "category": body.category,
        "description": body.description, "location": body.location,
        "used_on_asset": body.used_on_asset, "unit_cost": body.unit_cost,
        "reorder_level": body.reorder_level,
    }
    for field, new_value in detail_fields.items():
        if new_value is not None:
            old_value = getattr(part, field, None)
            if str(old_value) != str(new_value):
                setattr(part, field, new_value)
                field_changes.append(f"{field}: '{old_value}' -> '{new_value}'")

    if not field_changes:
        part.last_stock_take_at = datetime.utcnow()
        part.last_stock_take_by = current_user.name
        await db.flush()
        return {
            "success": True,
            "message": "No changes detected — part already matches the count.",
            "part": _part_summary(part),
            "qty_before": qty_before,
            "qty_after": qty_after,
            "changed": False,
        }

    part.updated_at = datetime.utcnow()
    part.last_stock_take_at = datetime.utcnow()
    part.last_stock_take_by = current_user.name

    change_summary = "; ".join(field_changes)
    full_reason = f"[STOCK TAKE] {body.reason or 'Physical count adjustment'} — {change_summary}"

    await _log_movement(
        db, part, "adjustment", abs(qty_after - qty_before),
        qty_before, qty_after, current_user,
        reason=full_reason, reference="Stock Take",
    )
    await db.flush()

    await ws_manager.broadcast_event("inventory", "inventory.adjustment", {
        "part_id":   str(part.id),
        "part_name": part.name,
        "qty_before": qty_before,
        "qty_after":  qty_after,
        "by":         current_user.name,
    })

    return {
        "success": True,
        "message": f"Stock take saved: {change_summary}",
        "part": _part_summary(part),
        "qty_before": qty_before,
        "qty_after": qty_after,
        "changed": True,
    }


# ── History: all movements ─────────────────────────────────────────────────

@router.get("/history")
async def get_history(
    part_id:     Optional[str]  = None,
    part_code:   Optional[str]  = None,
    movement:    Optional[str]  = None,   # stock_in | stock_out
    date_from:   Optional[str]  = None,
    date_to:     Optional[str]  = None,
    performed_by:Optional[str]  = None,
    limit:       int            = 200,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """Full stock movement history with optional filters."""
    await _ensure_table(db)

    conditions = []
    params: dict = {"limit": limit}

    if part_id:
        conditions.append("part_id = :part_id")
        params["part_id"] = part_id
    if part_code:
        conditions.append("part_code = :part_code")
        params["part_code"] = part_code
    if movement:
        conditions.append("movement = :movement")
        params["movement"] = movement
    if date_from:
        conditions.append("created_at >= :date_from")
        params["date_from"] = date_from
    if date_to:
        conditions.append("created_at <= :date_to")
        params["date_to"] = date_to + "T23:59:59"
    if performed_by:
        conditions.append("performed_by_name LIKE :performed_by")
        params["performed_by"] = f"%{performed_by}%"

    where = ("WHERE " + " AND ".join(conditions)) if conditions else ""
    sql   = f"""
        SELECT id, part_id, part_code, part_name, movement,
               quantity, qty_before, qty_after,
               reason, reference,
               performed_by_id, performed_by_name, created_at
        FROM stock_movements
        {where}
        ORDER BY created_at DESC
        LIMIT :limit
    """
    rows = (await db.execute(text(sql), params)).fetchall()
    return [dict(r._mapping) for r in rows]


@router.get("/history/{part_id}")
async def get_part_history(
    part_id: str,
    limit:   int = 50,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """Movement history for a single part."""
    await _ensure_table(db)
    rows = (await db.execute(
        text("""
            SELECT id, part_id, part_code, part_name, movement,
                   quantity, qty_before, qty_after,
                   reason, reference,
                   performed_by_id, performed_by_name, created_at
            FROM stock_movements
            WHERE part_id = :pid
            ORDER BY created_at DESC
            LIMIT :limit
        """),
        {"pid": part_id, "limit": limit}
    )).fetchall()
    return [dict(r._mapping) for r in rows]


# ── Export history to Excel ────────────────────────────────────────────────

@router.get("/history/export.xlsx")
async def export_history_excel(
    part_id:   Optional[str] = None,
    movement:  Optional[str] = None,
    date_from: Optional[str] = None,
    date_to:   Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """Export stock movement history to Excel."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        import io as _io
        from fastapi.responses import StreamingResponse as SR
    except ImportError:
        raise HTTPException(500, "openpyxl not installed")

    await _ensure_table(db)

    conditions, params = [], {"limit": 5000}
    if part_id:   conditions.append("part_id = :part_id");   params["part_id"] = part_id
    if movement:  conditions.append("movement = :movement");  params["movement"] = movement
    if date_from: conditions.append("created_at >= :df");     params["df"] = date_from
    if date_to:   conditions.append("created_at <= :dt");     params["dt"] = date_to + "T23:59:59"

    where = ("WHERE " + " AND ".join(conditions)) if conditions else ""
    rows  = (await db.execute(
        text(f"SELECT * FROM stock_movements {where} ORDER BY created_at DESC LIMIT :limit"),
        params
    )).fetchall()

    wb  = Workbook(); ws = wb.active; ws.title = "Stock History"
    thin = Side(style="thin", color="CCCCCC")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
    hf   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    hfl  = PatternFill("solid", start_color="1E3A5F")
    df   = Font(name="Arial", size=10)

    ws.merge_cells("A1:L1")
    ws["A1"] = f"CMMS Pro — Stock Movement History  |  {datetime.utcnow().strftime('%d %b %Y %H:%M')} UTC  |  {len(rows)} records"
    ws["A1"].font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    ws["A1"].fill = hfl
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 26

    headers = ["Date/Time","Part Code","Part Name","Movement","Qty Change","Qty Before","Qty After","Reason","Reference","Performed By"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font = hf; cell.fill = hfl; cell.border = bdr
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 22

    in_fill  = PatternFill("solid", start_color="DCFCE7")
    out_fill = PatternFill("solid", start_color="FEF9C4")

    for r, row in enumerate(rows, 3):
        m    = row._mapping
        fill = in_fill if "in" in (m["movement"] or "") else out_fill
        sign = "+" if "in" in (m["movement"] or "") else "−"
        row_data = [
            m["created_at"][:16].replace("T"," "),
            m["part_code"],
            m["part_name"],
            (m["movement"] or "").replace("_"," ").title(),
            f"{sign}{m['quantity']}",
            m["qty_before"],
            m["qty_after"],
            m["reason"]     or "",
            m["reference"]  or "",
            m["performed_by_name"] or "",
        ]
        for c, v in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = df; cell.fill = fill; cell.border = bdr
            cell.alignment = Alignment(vertical="center")
        ws.row_dimensions[r].height = 18

    for i, w in enumerate([18,12,28,14,12,12,12,24,18,18],1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A3"

    buf = _io.BytesIO(); wb.save(buf); buf.seek(0)
    fname = f"StockHistory_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
    from fastapi.responses import StreamingResponse
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'})